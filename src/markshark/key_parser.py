#!/usr/bin/env python3
"""
MarkShark Key Parser - Flexible answer key parsing for multiple formats.

Supports:
- Text files (.txt) - single or multi-version
- CSV files (.csv) - single or multi-column
- TSV files (.tsv) - tab-separated
- Excel files (.xlsx) - with optional Q# column

Header format (case-insensitive):
    ver:A code:101 default:3

    - ver: Version letter (e.g., A, B, C)
    - code: Test code for machine matching (e.g., 101, 202)
    - default: Points per question (default=1 if not specified)

    At least one of ver: or code: is required.

Answer formats:
    A           Single answer, default points
    A:3         Single answer, 3 points
    A^B         OR - either A or B gets full credit
    A^B:3       OR with custom points
    A@B         Partial credit (lenient) - +pts per correct, wrong ignored, spam=0
    A:2@B:2     Partial with custom points per answer
    A~B         Partial credit (strict) - +pts per correct, -pts per wrong, spam=0
    A:2~B:2     Strict partial with custom points
    *           Freebie - everyone gets default points
    *:3         Freebie with custom points
    (blank)     Discard question - remove from scoring denominator
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

# Optional Excel support
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ScoringMode(Enum):
    """How to score an answer."""
    SINGLE = "single"           # Single correct answer
    OR = "or"                   # Multiple acceptable answers (any one is full credit)
    AND = "and"                 # Must select ALL correct answers (exact match required)
    PARTIAL_LENIENT = "lenient" # +pts per correct, wrong ignored, spam=0
    PARTIAL_STRICT = "strict"   # +pts per correct, -pts per wrong, spam=0
    FREEBIE = "freebie"         # Everyone gets points
    DISCARD = "discard"         # Remove from scoring


@dataclass
class AnswerSpec:
    """Specification for a single question's answer."""
    mode: ScoringMode
    correct_answers: Dict[str, float] = field(default_factory=dict)  # answer -> points
    total_points: float = 1.0  # Max points for this question

    @property
    def is_scorable(self) -> bool:
        """Whether this question contributes to the score denominator."""
        return self.mode != ScoringMode.DISCARD


@dataclass
class VersionKey:
    """Answer key for a single version/code."""
    version: Optional[str] = None       # e.g., "A", "B"
    code: Optional[str] = None          # e.g., "101", "202"
    default_points: float = 1.0         # Default points per question
    answers: List[AnswerSpec] = field(default_factory=list)

    @property
    def identifier(self) -> str:
        """Primary identifier for this key (version preferred, then code)."""
        return self.version or self.code or "A"

    @property
    def num_questions(self) -> int:
        """Number of questions in this key."""
        return len(self.answers)

    @property
    def max_points(self) -> float:
        """Maximum possible points for this version."""
        return sum(a.total_points for a in self.answers if a.is_scorable)

    @property
    def num_scorable(self) -> int:
        """Number of questions that count toward score."""
        return sum(1 for a in self.answers if a.is_scorable)


@dataclass
class AnswerKeySet:
    """Collection of answer keys for an exam (possibly multi-version)."""
    keys: Dict[str, VersionKey] = field(default_factory=dict)  # identifier -> VersionKey
    code_to_version: Dict[str, str] = field(default_factory=dict)  # code -> version/identifier mapping

    def get_key(self, version: Optional[str] = None, code: Optional[str] = None) -> Optional[VersionKey]:
        """
        Get the appropriate key for a student's version/code.

        Priority:
        1. Exact version match
        2. Code lookup via code_to_version mapping
        3. Code as direct key identifier
        4. First available key (fallback)
        """
        # Try version first
        if version and version.upper() in self.keys:
            return self.keys[version.upper()]

        # Try code lookup via mapping
        if code and code in self.code_to_version:
            identifier = self.code_to_version[code]
            if identifier in self.keys:
                return self.keys[identifier]

        # Try code as direct key identifier
        if code and code in self.keys:
            return self.keys[code]

        # Try finding a key by its code property
        if code:
            for key in self.keys.values():
                if key.code == code:
                    return key

        # Fallback to first key
        if self.keys:
            return next(iter(self.keys.values()))

        return None

    @property
    def is_multi_version(self) -> bool:
        """Whether this key set has multiple versions."""
        return len(self.keys) > 1

    @property
    def has_codes(self) -> bool:
        """Whether any keys have test codes."""
        return bool(self.code_to_version)

    @property
    def versions(self) -> List[str]:
        """List of version identifiers."""
        return sorted(self.keys.keys())


# =============================================================================
# Parsing Functions
# =============================================================================

def parse_header(header_text: str) -> Tuple[Optional[str], Optional[str], float]:
    """
    Parse a key header line to extract version, code, and default points.

    Examples:
        "ver:A" -> ("A", None, 1.0)
        "code:101" -> (None, "101", 1.0)
        "ver:B code:202 default:3" -> ("B", "202", 3.0)
        "VER:A DEFAULT:2" -> ("A", None, 2.0)  # case-insensitive

    Returns:
        (version, code, default_points)

    Raises:
        ValueError if neither ver: nor code: is found
    """
    text = header_text.strip()

    version = None
    code = None
    default_points = 1.0

    # Case-insensitive regex patterns
    ver_match = re.search(r'\bver(?:sion)?:\s*([A-Za-z0-9]+)', text, re.IGNORECASE)
    code_match = re.search(r'\bcode:\s*([A-Za-z0-9]+)', text, re.IGNORECASE)
    default_match = re.search(r'\bdefault(?:points)?:\s*([0-9.]+)', text, re.IGNORECASE)

    if ver_match:
        version = ver_match.group(1).upper()

    if code_match:
        code = code_match.group(1)

    if default_match:
        try:
            default_points = float(default_match.group(1))
        except ValueError:
            pass

    # Require at least one identifier
    if version is None and code is None:
        raise ValueError(f"Header must contain 'ver:' or 'code:': {header_text}")

    return version, code, default_points


def parse_answer(answer_text: str, default_points: float = 1.0) -> AnswerSpec:
    """
    Parse a single answer specification.

    Examples:
        "A" -> single answer A, default points
        "A:3" -> single answer A, 3 points
        "A^B" -> A or B accepted (OR), default points
        "A^B:3" -> A or B accepted, 3 points
        "A&B" -> must select BOTH A and B (AND), default points
        "A&B:4" -> must select both A and B, 4 points
        "A@B" -> partial lenient, A=default, B=default
        "A:2@B:1" -> partial lenient, A=2pts, B=1pt
        "A~B" -> partial strict, A=default, B=default
        "A:2~B:1" -> partial strict, A=2pts, B=1pt
        "*" -> freebie, default points
        "*:3" -> freebie, 3 points
        "" or " " -> discard question

    Returns:
        AnswerSpec with parsed configuration
    """
    text = answer_text.strip().upper()

    # Empty = discard
    if not text:
        return AnswerSpec(mode=ScoringMode.DISCARD, total_points=0.0)

    # Freebie
    if text.startswith("*"):
        points = default_points
        if ":" in text:
            try:
                points = float(text.split(":")[1])
            except (ValueError, IndexError):
                pass
        return AnswerSpec(
            mode=ScoringMode.FREEBIE,
            correct_answers={},
            total_points=points,
        )

    # Check for operators (in order of precedence)
    # @ = partial lenient, ~ = partial strict, & = AND (exact match), ^ = OR

    if "@" in text:
        # Partial credit (lenient)
        return _parse_partial(text, "@", ScoringMode.PARTIAL_LENIENT, default_points)

    if "~" in text:
        # Partial credit (strict)
        return _parse_partial(text, "~", ScoringMode.PARTIAL_STRICT, default_points)

    if "&" in text:
        # AND mode - must select ALL specified answers (exact match)
        parts = text.split("&")
        total_points = default_points
        answers = {}

        for part in parts:
            part = part.strip()
            if ":" in part:
                letter, pts_str = part.split(":", 1)
                letter = letter.strip()
                try:
                    # In AND mode, the points on the last part sets the total
                    total_points = float(pts_str.strip())
                except ValueError:
                    pass
                if letter and letter.isalpha():
                    answers[letter] = total_points
            elif part and part.isalpha():
                answers[part] = total_points

        return AnswerSpec(
            mode=ScoringMode.AND,
            correct_answers=answers,
            total_points=total_points,
        )

    if "^" in text:
        # OR mode
        parts = text.split("^")
        # Check if last part has points override for the whole thing
        total_points = default_points
        answers = {}

        for part in parts:
            part = part.strip()
            if ":" in part:
                letter, pts_str = part.split(":", 1)
                letter = letter.strip()
                try:
                    # In OR mode, the points on any answer set the total
                    total_points = float(pts_str.strip())
                except ValueError:
                    pass
                if letter and letter.isalpha():
                    answers[letter] = total_points
            elif part and part.isalpha():
                answers[part] = total_points

        return AnswerSpec(
            mode=ScoringMode.OR,
            correct_answers=answers,
            total_points=total_points,
        )

    # Single answer
    if ":" in text:
        letter, pts_str = text.split(":", 1)
        letter = letter.strip()
        try:
            points = float(pts_str.strip())
        except ValueError:
            points = default_points
    else:
        letter = text.strip()
        points = default_points

    if not letter or not letter[0].isalpha():
        # Invalid answer, treat as discard
        return AnswerSpec(mode=ScoringMode.DISCARD, total_points=0.0)

    return AnswerSpec(
        mode=ScoringMode.SINGLE,
        correct_answers={letter[0]: points},
        total_points=points,
    )


def _parse_partial(text: str, operator: str, mode: ScoringMode, default_points: float) -> AnswerSpec:
    """Parse partial credit answer (A@B or A~B style)."""
    parts = text.split(operator)
    answers = {}
    total_points = 0.0

    for part in parts:
        part = part.strip()
        if ":" in part:
            letter, pts_str = part.split(":", 1)
            letter = letter.strip()
            try:
                pts = float(pts_str.strip())
            except ValueError:
                pts = default_points
        else:
            letter = part
            pts = default_points

        if letter and letter[0].isalpha():
            answers[letter[0]] = pts
            total_points += pts

    return AnswerSpec(
        mode=mode,
        correct_answers=answers,
        total_points=total_points,
    )


# =============================================================================
# File Loading Functions
# =============================================================================

def load_key_file(path: Union[str, Path]) -> AnswerKeySet:
    """
    Load answer key(s) from a file, auto-detecting format.

    Supported formats:
    - .txt - Text file (single or multi-version)
    - .csv - Comma-separated values
    - .tsv - Tab-separated values
    - .xlsx - Excel workbook

    Returns:
        AnswerKeySet containing all parsed keys
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required to read Excel files. Install with: pip install openpyxl")
        return _load_xlsx(path)
    elif suffix == ".tsv":
        return _load_delimited(path, delimiter="\t")
    elif suffix == ".csv":
        return _load_delimited(path, delimiter=",")
    else:
        # Default to text file parsing
        return _load_text(path)


def _read_text_permissive(path: Path) -> str:
    """Read a text file regardless of encoding or newline style.

    Teachers often save answer keys from Word, Notepad, or TextEdit, which
    may produce UTF-8, UTF-8 with BOM, Windows-1252, or Mac-Roman depending
    on which dialog option they pick.  Python's text mode already normalises
    all newline styles (\\r\\n, \\r, \\n) to \\n, so the only real issue is
    encoding.

    Strategy: try UTF-8 first (handles BOM automatically via utf-8-sig),
    fall back to Latin-1 which accepts every possible byte value — so it
    never raises an error.  For answer key files (A-E letters, digits,
    punctuation) the encoding rarely matters, but this prevents cryptic
    UnicodeDecodeError crashes when a teacher saves from Word with the
    wrong encoding selected.
    """
    raw = path.read_bytes()

    # Strip UTF-8 BOM if present (Word on Windows adds this)
    if raw.startswith(b"\xef\xbb\xbf"):
        raw = raw[3:]

    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        # Latin-1 maps every byte 0x00-0xFF to a character, so this
        # never fails.  For answer keys (letters, digits, operators)
        # the result is identical to the original intent.
        return raw.decode("latin-1")


def _load_text(path: Path) -> AnswerKeySet:
    """Load a text-format answer key file."""
    lines = _read_text_permissive(path).splitlines(keepends=True)

    keys: Dict[str, VersionKey] = {}
    code_to_version: Dict[str, str] = {}

    current_key: Optional[VersionKey] = None

    for line_num, line in enumerate(lines, 1):
        line = line.strip()

        # Skip blank lines
        if not line:
            continue

        # Skip comment lines (# followed by space or ##)
        if line.startswith("##") or (line.startswith("#") and len(line) > 1 and line[1] == " "):
            continue

        # Check for header line (contains ver: or code:)
        if re.search(r'\b(ver|version|code):', line, re.IGNORECASE):
            try:
                version, code, default_points = parse_header(line)
                current_key = VersionKey(
                    version=version,
                    code=code,
                    default_points=default_points,
                )

                # Store the key
                identifier = version or code
                keys[identifier] = current_key

                # Map code to the identifier (version if both exist, code if code-only)
                if code:
                    code_to_version[code] = identifier

            except ValueError as e:
                raise ValueError(f"Line {line_num}: {e}")
            continue

        # If no current key, this is an error
        if current_key is None:
            raise ValueError(f"Line {line_num}: Answer found before header line. "
                           f"Expected 'ver:X' or 'code:X' header first.")

        # Parse answer line(s)
        # Could be comma-separated on one line or one per line
        if "," in line:
            # Comma-separated answers on one line
            for answer_text in line.split(","):
                answer_text = answer_text.strip()
                if answer_text:
                    spec = parse_answer(answer_text, current_key.default_points)
                    current_key.answers.append(spec)
        else:
            # Single answer
            spec = parse_answer(line, current_key.default_points)
            current_key.answers.append(spec)

    if not keys:
        raise ValueError("No valid answer keys found in file")

    return AnswerKeySet(keys=keys, code_to_version=code_to_version)


def _load_delimited(path: Path, delimiter: str = ",") -> AnswerKeySet:
    """Load a CSV or TSV answer key file."""
    # Use the same permissive encoding strategy as _load_text so CSV keys
    # saved from Excel or Word with non-UTF-8 encoding don't crash.
    import io
    text = _read_text_permissive(path)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)

    if not rows:
        raise ValueError("Empty file")

    return _parse_tabular_data(rows)


def _load_xlsx(path: Path) -> AnswerKeySet:
    """Load an Excel answer key file."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Find the key sheet (skip Instructions if present)
    sheet_names = wb.sheetnames
    key_sheet = None

    for name in sheet_names:
        if name.lower() not in ("instructions", "help", "readme"):
            key_sheet = wb[name]
            break

    if key_sheet is None:
        key_sheet = wb.active

    # Read all rows
    rows = []
    for row in key_sheet.iter_rows(values_only=True):
        # Convert None to empty string
        rows.append([str(cell) if cell is not None else "" for cell in row])

    wb.close()

    if not rows:
        raise ValueError("Empty worksheet")

    return _parse_tabular_data(rows)


def _parse_tabular_data(rows: List[List[str]]) -> AnswerKeySet:
    """
    Parse tabular data (from CSV, TSV, or Excel).

    Expects:
    - First row with headers: optional Q# column, then ver:A, ver:B, etc.
    - Subsequent rows: optional Q number, then answers for each version
    """
    if not rows:
        raise ValueError("No data rows")

    # Find the header row (first row containing ver: or code:)
    header_row_idx = None
    for idx, row in enumerate(rows):
        row_text = " ".join(str(cell) for cell in row)
        if re.search(r'\b(ver|version|code):', row_text, re.IGNORECASE):
            header_row_idx = idx
            break

    if header_row_idx is None:
        raise ValueError("No header row found. Expected row with 'ver:X' or 'code:X' columns.")

    header_row = rows[header_row_idx]

    # Parse header columns to find version columns
    version_columns: List[Tuple[int, str, str, float]] = []  # (col_idx, version, code, default)
    q_column_idx = None

    for col_idx, cell in enumerate(header_row):
        cell_text = str(cell).strip()

        # Check for Q# column
        if cell_text.upper() in ("Q#", "Q", "QUESTION", "#", "NUM", "NUMBER"):
            q_column_idx = col_idx
            continue

        # Check for version/code header
        if re.search(r'\b(ver|version|code):', cell_text, re.IGNORECASE):
            try:
                version, code, default_points = parse_header(cell_text)
                version_columns.append((col_idx, version, code, default_points))
            except ValueError:
                pass

    if not version_columns:
        raise ValueError("No version columns found in header row")

    # Create VersionKey objects
    keys: Dict[str, VersionKey] = {}
    code_to_version: Dict[str, str] = {}

    for col_idx, version, code, default_points in version_columns:
        identifier = version or code
        keys[identifier] = VersionKey(
            version=version,
            code=code,
            default_points=default_points,
        )
        # Map code to the identifier (version if both exist, code if code-only)
        if code:
            code_to_version[code] = identifier

    # Parse answer rows
    for row_idx in range(header_row_idx + 1, len(rows)):
        row = rows[row_idx]

        # Skip empty rows
        if not any(str(cell).strip() for cell in row):
            continue

        # Skip comment/note rows (start with #, (, or common note prefixes)
        first_cell = str(row[0]).strip().lower() if row else ""
        if first_cell.startswith("#") or first_cell.startswith("("):
            continue
        if first_cell.startswith("note") or first_cell.startswith("example"):
            continue

        # Skip rows where the first version column is empty or contains non-answer text
        # This catches rows like "(Examples above...)" that span multiple columns
        first_version_col = version_columns[0][0] if version_columns else 1
        if first_version_col < len(row):
            first_answer = str(row[first_version_col]).strip()
            # If the answer cell is empty or looks like a note, skip this row
            if not first_answer or first_answer.startswith("(") or len(first_answer) > 20:
                continue

        # Parse answers for each version
        for col_idx, version, code, default_points in version_columns:
            if col_idx < len(row):
                answer_text = str(row[col_idx]).strip()
                identifier = version or code
                spec = parse_answer(answer_text, keys[identifier].default_points)
                keys[identifier].answers.append(spec)

    # Validate all versions have same number of questions
    num_questions = [len(k.answers) for k in keys.values()]
    if len(set(num_questions)) > 1:
        print(f"Warning: Version keys have different numbers of questions: {num_questions}")

    return AnswerKeySet(keys=keys, code_to_version=code_to_version)


# =============================================================================
# Scoring Functions
# =============================================================================

def score_answer(student_answer: Optional[str], answer_spec: AnswerSpec) -> Tuple[float, str]:
    """
    Score a student's answer against an answer specification.

    Args:
        student_answer: What the student bubbled (e.g., "A", "B,C", None for blank)
        answer_spec: The answer specification for this question

    Returns:
        (points_earned, status) where status is one of:
        - "correct" - full credit
        - "partial" - partial credit (for @ or ~ modes)
        - "incorrect" - wrong answer
        - "blank" - no answer given
        - "multi" - multiple bubbles filled
        - "spam" - too many bubbles (anti-spam triggered)
        - "freebie" - everyone gets credit
        - "discard" - question not scored
    """
    # Handle discard questions
    if answer_spec.mode == ScoringMode.DISCARD:
        return 0.0, "discard"

    # Handle freebie questions
    if answer_spec.mode == ScoringMode.FREEBIE:
        return answer_spec.total_points, "freebie"

    # Handle blank answers
    if student_answer is None or student_answer.strip() == "":
        return 0.0, "blank"

    # Parse student's answer(s)
    student_answer = student_answer.strip().upper()

    # Check for multiple bubbles
    if "," in student_answer:
        student_choices = set(a.strip() for a in student_answer.split(",") if a.strip())
    else:
        student_choices = {student_answer}

    # For single answer mode
    if answer_spec.mode == ScoringMode.SINGLE:
        correct_letter = list(answer_spec.correct_answers.keys())[0]
        if len(student_choices) > 1:
            return 0.0, "multi"
        if correct_letter in student_choices:
            return answer_spec.total_points, "correct"
        return 0.0, "incorrect"

    # For OR mode (any correct answer gets full credit)
    if answer_spec.mode == ScoringMode.OR:
        if len(student_choices) > 1:
            return 0.0, "multi"
        for choice in student_choices:
            if choice in answer_spec.correct_answers:
                return answer_spec.total_points, "correct"
        return 0.0, "incorrect"

    # For AND mode (must select ALL correct answers exactly)
    if answer_spec.mode == ScoringMode.AND:
        correct_set = set(answer_spec.correct_answers.keys())
        # Must match exactly - no more, no less
        if student_choices == correct_set:
            return answer_spec.total_points, "correct"
        # Any deviation is incorrect (no partial credit)
        return 0.0, "incorrect"

    # For partial credit modes
    if answer_spec.mode in (ScoringMode.PARTIAL_LENIENT, ScoringMode.PARTIAL_STRICT):
        num_correct = len(answer_spec.correct_answers)

        # Check for spam (more answers than correct options)
        if len(student_choices) > num_correct:
            return 0.0, "spam"

        # Calculate points
        points = 0.0
        correct_count = 0
        wrong_count = 0

        for choice in student_choices:
            if choice in answer_spec.correct_answers:
                points += answer_spec.correct_answers[choice]
                correct_count += 1
            else:
                wrong_count += 1
                if answer_spec.mode == ScoringMode.PARTIAL_STRICT:
                    # Subtract points for wrong answers in strict mode
                    # Use average correct answer points as penalty
                    avg_points = sum(answer_spec.correct_answers.values()) / len(answer_spec.correct_answers)
                    points -= avg_points

        # Floor at 0
        points = max(0.0, points)

        # Determine status
        if correct_count == num_correct and wrong_count == 0:
            return points, "correct"
        elif correct_count > 0:
            return points, "partial"
        else:
            return 0.0, "incorrect"

    # Fallback (shouldn't reach here)
    return 0.0, "incorrect"


def score_student(
    student_answers: List[Optional[str]],
    version_key: VersionKey
) -> Tuple[float, float, Dict[str, int]]:
    """
    Score a student's complete answer sheet.

    Args:
        student_answers: List of student's answers (one per question)
        version_key: The answer key for this version

    Returns:
        (points_earned, max_points, status_counts)

        status_counts is a dict with counts of each status type:
        {"correct": n, "incorrect": n, "blank": n, "multi": n, "partial": n, ...}
    """
    points_earned = 0.0
    max_points = 0.0
    status_counts: Dict[str, int] = {}

    for q_idx, answer_spec in enumerate(version_key.answers):
        student_ans = student_answers[q_idx] if q_idx < len(student_answers) else None

        pts, status = score_answer(student_ans, answer_spec)
        points_earned += pts

        # Only add to max_points for scorable questions
        if answer_spec.is_scorable:
            max_points += answer_spec.total_points

        status_counts[status] = status_counts.get(status, 0) + 1

    return points_earned, max_points, status_counts


# =============================================================================
# Serialization Functions
# =============================================================================

def answer_spec_to_text(spec: AnswerSpec, default_points: float = 1.0) -> str:
    """
    Serialize an AnswerSpec back to its canonical text representation.

    Point values are included only when they differ from *default_points*
    so that round-tripping through parse_answer → answer_spec_to_text is
    clean.

    Examples (default_points=1.0):
        SINGLE  A, 1pt  -> "A"
        SINGLE  A, 3pt  -> "A:3"
        OR      A^B     -> "A^B"
        AND     A&B     -> "A&B"
        PARTIAL A:2@B:1 -> "A:2@B:1"
        FREEBIE 1pt     -> "*"
        FREEBIE 5pt     -> "*:5"
        DISCARD         -> ""
    """
    if spec.mode == ScoringMode.DISCARD:
        return ""

    if spec.mode == ScoringMode.FREEBIE:
        if spec.total_points != default_points:
            return f"*:{spec.total_points:g}"
        return "*"

    keys_sorted = sorted(spec.correct_answers.keys())
    if not keys_sorted:
        return ""

    # Choose the operator character for the mode
    op_map = {
        ScoringMode.OR: "^",
        ScoringMode.AND: "&",
        ScoringMode.PARTIAL_LENIENT: "@",
        ScoringMode.PARTIAL_STRICT: "~",
    }
    operator = op_map.get(spec.mode)

    if operator is not None:
        # Multi-answer modes — include per-answer points when non-default
        all_default = all(
            spec.correct_answers[k] == default_points for k in keys_sorted
        )
        if all_default:
            text = operator.join(keys_sorted)
            # Append total override only for OR/AND when total != default
            if spec.mode in (ScoringMode.OR, ScoringMode.AND):
                if spec.total_points != default_points:
                    text += f":{spec.total_points:g}"
            return text
        else:
            # Per-answer point values
            parts = [
                f"{k}:{spec.correct_answers[k]:g}" for k in keys_sorted
            ]
            return operator.join(parts)

    # SINGLE mode
    letter = keys_sorted[0]
    if spec.total_points != default_points:
        return f"{letter}:{spec.total_points:g}"
    return letter


def write_key_file(
    key_set: AnswerKeySet, path: Union[str, Path], fmt: str = "txt"
) -> None:
    """
    Write an AnswerKeySet to a file.

    Args:
        key_set: The answer key set to serialize.
        path:    Destination file path.
        fmt:     ``"txt"`` for the text format that ``_load_text`` reads,
                 ``"xlsx"`` for an Excel workbook with an *Answer Key* sheet.
    """
    path = Path(path)
    if fmt == "xlsx":
        _write_xlsx(key_set, path)
    else:
        _write_text(key_set, path)


def _write_text(key_set: AnswerKeySet, path: Path) -> None:
    """Write key set in the ver:/code:/default: text format."""
    lines: List[str] = []
    for identifier in sorted(key_set.keys):
        vk = key_set.keys[identifier]
        # Build header
        parts: List[str] = []
        if vk.version:
            parts.append(f"ver:{vk.version}")
        if vk.code:
            parts.append(f"code:{vk.code}")
        if vk.default_points != 1.0:
            parts.append(f"default:{vk.default_points:g}")
        lines.append(" ".join(parts))

        # Build comma-separated answer line
        answer_texts = [
            answer_spec_to_text(spec, vk.default_points) for spec in vk.answers
        ]
        lines.append(",".join(answer_texts))

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_xlsx(key_set: AnswerKeySet, path: Path) -> None:
    """Write key set as an Excel workbook."""
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required to write Excel files. "
            "Install with: pip install openpyxl"
        )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Answer Key"

    # Build column headers: Q# then one column per version
    sorted_ids = sorted(key_set.keys)
    headers = ["Q#"]
    for vid in sorted_ids:
        vk = key_set.keys[vid]
        parts: List[str] = []
        if vk.version:
            parts.append(f"ver:{vk.version}")
        if vk.code:
            parts.append(f"code:{vk.code}")
        if vk.default_points != 1.0:
            parts.append(f"default:{vk.default_points:g}")
        headers.append(" ".join(parts))
    ws.append(headers)

    # Determine max question count across versions
    max_q = max((vk.num_questions for vk in key_set.keys.values()), default=0)

    for q_idx in range(max_q):
        row = [q_idx + 1]
        for vid in sorted_ids:
            vk = key_set.keys[vid]
            if q_idx < len(vk.answers):
                row.append(answer_spec_to_text(vk.answers[q_idx], vk.default_points))
            else:
                row.append("")
        ws.append(row)

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 8)

    wb.save(path)


# =============================================================================
# Legacy Compatibility Functions
# =============================================================================

def to_legacy_keys_dict(key_set: AnswerKeySet) -> Dict[str, List[str]]:
    """
    Convert AnswerKeySet to legacy format (Dict[version, List[str]]).

    Used for backward compatibility with existing scoring code.
    Compound answers are reconstructed with their separator:
      AND -> "B&C",  OR -> "B^C",
      PARTIAL_LENIENT -> "A@B",  PARTIAL_STRICT -> "A~B"
    """
    result = {}

    for identifier, version_key in key_set.keys.items():
        letters = []
        for spec in version_key.answers:
            if spec.mode == ScoringMode.DISCARD:
                letters.append("")
            elif spec.mode == ScoringMode.FREEBIE:
                letters.append("*")
            elif spec.correct_answers:
                # Reconstruct the separator-delimited key string
                # so downstream code (annotation, CSV KEY row) sees
                # the full compound answer, not just the first letter.
                keys_sorted = sorted(spec.correct_answers.keys())
                if spec.mode == ScoringMode.AND:
                    letters.append("&".join(keys_sorted))
                elif spec.mode == ScoringMode.OR:
                    letters.append("^".join(keys_sorted))
                elif spec.mode == ScoringMode.PARTIAL_LENIENT:
                    letters.append("@".join(keys_sorted))
                elif spec.mode == ScoringMode.PARTIAL_STRICT:
                    letters.append("~".join(keys_sorted))
                else:
                    # SINGLE mode — just the one letter
                    letters.append(keys_sorted[0])
            else:
                letters.append("")
        result[identifier] = letters

    return result


