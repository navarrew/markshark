#!/usr/bin/env python3
"""
Generate the Answer Key Excel template with Instructions tab.

Run this script to create/update the template file:
    python -m markshark.assets.answer_key_template
"""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl is required. Install with: pip install openpyxl")


def create_answer_key_template(output_path: Path = None) -> Path:
    """
    Create an Excel answer key template with Instructions and Key tabs.

    Returns the path to the created file.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required to create Excel templates")

    if output_path is None:
        output_path = Path(__file__).parent / "answer_key_template.xlsx"

    wb = Workbook()

    # =========================================================================
    # Instructions Tab
    # =========================================================================
    instructions = wb.active
    instructions.title = "Instructions"

    # Styles
    title_font = Font(bold=True, size=16, color="1F4E79")
    header_font = Font(bold=True, size=12, color="1F4E79")
    code_font = Font(name="Consolas", size=11)
    example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    row = 1

    # Title
    instructions.cell(row=row, column=1, value="MarkShark Answer Key Format Guide").font = title_font
    row += 2

    # Header format section
    instructions.cell(row=row, column=1, value="HEADER FORMAT").font = header_font
    row += 1
    instructions.cell(row=row, column=1, value="Each column header identifies the version/test code. At least one of ver: or code: is required.")
    row += 2

    headers_data = [
        ("Header", "Meaning"),
        ("ver:A", "Version A"),
        ("ver:B default:3", "Version B, 3 points per question"),
        ("code:101", "Test code 101"),
        ("ver:A code:101 default:2", "Version A with code 101, 2 points per question"),
    ]
    for i, (header, meaning) in enumerate(headers_data):
        c1 = instructions.cell(row=row, column=1, value=header)
        c2 = instructions.cell(row=row, column=2, value=meaning)
        if i == 0:
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
        else:
            c1.font = code_font
            c1.fill = example_fill
        row += 1

    row += 1

    # Answer format section
    instructions.cell(row=row, column=1, value="ANSWER FORMATS").font = header_font
    row += 1
    instructions.cell(row=row, column=1, value="Use these formats in the answer cells:")
    row += 2

    answer_data = [
        ("Format", "Meaning", "Points"),
        ("A", "Single answer A", "default"),
        ("A:3", "Single answer A", "3 points"),
        ("A^B", "A OR B accepted (either gets full credit)", "default"),
        ("A^B:4", "A OR B accepted", "4 points"),
        ("A&B", "A AND B required (must select both, exact match)", "default"),
        ("A&B:4", "A AND B required", "4 points"),
        ("A@B", "Partial credit (lenient): +pts per correct, wrong ignored", "sum of parts"),
        ("A:2@B:1", "Partial with custom points: A=2pts, B=1pt", "A=2, B=1"),
        ("A~B", "Partial credit (strict): +pts per correct, -pts per wrong", "sum of parts"),
        ("*", "Freebie - everyone gets points", "default"),
        ("*:5", "Freebie", "5 points"),
        ("(blank)", "Discard question - remove from scoring", "0"),
    ]
    for i, row_data in enumerate(answer_data):
        for j, val in enumerate(row_data):
            c = instructions.cell(row=row, column=j+1, value=val)
            if i == 0:
                c.font = Font(bold=True)
            elif j == 0:
                c.font = code_font
                c.fill = example_fill
        row += 1

    row += 1

    # Partial credit warning
    instructions.cell(row=row, column=1, value="IMPORTANT: Partial Credit Questions").font = header_font
    row += 1

    warning_text = (
        "When using partial credit (@ or ~), your question MUST tell students how many answers to select!\n"
        "Example: \"Select the TWO correct answers\" NOT \"Select all that apply\"\n\n"
        "Anti-spam rule: If a student fills more bubbles than there are correct answers, they get ZERO points."
    )
    cell = instructions.cell(row=row, column=1, value=warning_text)
    cell.alignment = Alignment(wrap_text=True)
    instructions.merge_cells(start_row=row, start_column=1, end_row=row+3, end_column=3)
    row += 5

    # Partial credit examples
    instructions.cell(row=row, column=1, value="Partial Credit Scoring Examples:").font = Font(bold=True)
    row += 1

    partial_examples = [
        ("Question", "Key", "Student Answers", "Points"),
        ("Select TWO correct (B and C)", "B@C", "B, C", "2 pts (full credit)"),
        ("Select TWO correct (B and C)", "B@C", "B only", "1 pt (partial)"),
        ("Select TWO correct (B and C)", "B@C", "A, C", "1 pt (lenient: wrong ignored)"),
        ("Select TWO correct (B and C)", "B~C", "A, C", "0 pts (strict: wrong subtracts)"),
        ("Select TWO correct (B and C)", "B@C", "A, B, C", "0 pts (spam: 3 > 2 correct)"),
    ]
    for i, row_data in enumerate(partial_examples):
        for j, val in enumerate(row_data):
            c = instructions.cell(row=row, column=j+1, value=val)
            if i == 0:
                c.font = Font(bold=True)
        row += 1

    row += 1

    # Q# column info
    instructions.cell(row=row, column=1, value="Q# COLUMN").font = header_font
    row += 1
    instructions.cell(row=row, column=1, value="The Q# column is optional and for your reference only. MarkShark ignores it and uses row position.")
    row += 2

    # Set column widths for Instructions
    instructions.column_dimensions['A'].width = 40
    instructions.column_dimensions['B'].width = 45
    instructions.column_dimensions['C'].width = 25
    instructions.column_dimensions['D'].width = 25

    # =========================================================================
    # Answer Key Tab
    # =========================================================================
    key_sheet = wb.create_sheet("Answer Key")

    # Styles for key sheet
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    qnum_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row
    headers = ["Q#", "ver:A", "ver:B", "ver:C", "ver:D"]
    for col, header in enumerate(headers, 1):
        cell = key_sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Pre-populate 100 rows with Q# formula
    for q in range(1, 101):
        row_num = q + 1

        # Q# column with formula (shows Q# only if there's content in column B)
        qnum_cell = key_sheet.cell(row=row_num, column=1)
        qnum_cell.value = f'=IF(B{row_num}="","",{q})'
        qnum_cell.fill = qnum_fill
        qnum_cell.border = thin_border
        qnum_cell.alignment = Alignment(horizontal='center')

        # Answer columns (empty but bordered)
        for col in range(2, 6):
            cell = key_sheet.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # Set column widths
    key_sheet.column_dimensions['A'].width = 8
    for col in range(2, 6):
        key_sheet.column_dimensions[get_column_letter(col)].width = 20

    # Freeze header row
    key_sheet.freeze_panes = 'A2'

    # Add some example data in first few rows (user can delete)
    example_answers = [
        ["A", "B", "C", "D"],
        ["B:3", "A:3", "D:3", "C:3"],
        ["A^B", "B^C", "C^D", "A^D"],
        ["A@B", "B@C", "C@D", "A@D"],
        ["*", "*", "*:2", "*"],
    ]
    for q, answers in enumerate(example_answers, 2):
        for col, ans in enumerate(answers, 2):
            key_sheet.cell(row=q, column=col, value=ans)

    # Add note about examples
    note_row = 8
    key_sheet.cell(row=note_row, column=1, value="(Examples above - replace with your answers)")
    key_sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    note_cell = key_sheet.cell(row=note_row, column=1)
    note_cell.font = Font(italic=True, color="808080")

    # Save workbook
    wb.save(output_path)
    print(f"Created template: {output_path}")

    return output_path


if __name__ == "__main__":
    create_answer_key_template()
