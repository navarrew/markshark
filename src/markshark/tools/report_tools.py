#!/usr/bin/env python3
"""
MarkShark
report_tools.py
Generate teacher-friendly Excel reports from scored CSV results

Features:
- Multi-version support: separate tab per version
- Roster matching: flags absent students and orphan scans
- Color-coded item quality indicators
- Summary statistics and item analysis
"""

from __future__ import annotations
from typing import Optional, List, Dict, Tuple
import os

import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel report generation. "
        "Install it with: pip install openpyxl"
    )

try:
    from rapidfuzz import fuzz
except ImportError:
    raise ImportError(
        "rapidfuzz is required for student roster matching. "
        "Install it with: pip install rapidfuzz"
    )

from .stats_tools import (
    detect_item_columns,
    detect_key_row_index,
    prepare_correctness_matrix,
    point_biserial,
    kr20,
    kr21,
)


# ==================== CORRECTIONS HANDLING ====================

def _normalize_id(val) -> str:
    """Normalize a student ID for comparison: stringify, strip, remove .0 suffix."""
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def _find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """Find the first column name from candidates that exists in df."""
    for col in candidates:
        if col in df.columns:
            return col
    return None


def _clean_id_string(val) -> str:
    """
    Clean an ID value that may have been read as a number by Excel/pandas.

    Handles: 1234567890.0 -> "1234567890", NaN -> "", etc.
    """
    if pd.isna(val):
        return ''
    s = str(val).strip()
    # Remove .0 suffix from floats that were integers
    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]
    if s.lower() in ('nan', 'none', ''):
        return ''
    return s


def _load_corrections_csv(corrections_path: str) -> pd.DataFrame:
    """
    Load corrections from the new append-only CSV format.

    Expected format::

        # applies to: /path/to/results.csv
        timestamp,type,student_id,field,original,corrected,reason

    Only *effective* corrections are returned (REVERTs cancel earlier entries).

    Returns:
        DataFrame with columns: student_id, question, corrected_answer, original_answer
    """
    import sys

    rows: list[dict] = []
    with open(corrections_path, "r", newline="", encoding="utf-8") as f:
        first_line = f.readline()
        # Skip the "# applies to:" header comment
        if not (first_line.startswith("# applies to: ")
                or first_line.startswith("# applies_to: ")):
            f.seek(0)

        import csv as csv_mod
        reader = csv_mod.DictReader(f)
        if reader.fieldnames is None:
            return pd.DataFrame(columns=['student_id', 'question', 'corrected_answer', 'original_answer'])
        for row in reader:
            rows.append(row)

    if not rows:
        return pd.DataFrame(columns=['student_id', 'question', 'corrected_answer', 'original_answer'])

    # Compute effective corrections: apply entries in order, REVERTs remove.
    # Store both corrected AND original values.
    effective: dict[str, dict[str, dict[str, str]]] = {}  # student_id -> {field: {corrected, original}}
    for row in rows:
        sid = _clean_id_string(row.get('student_id', ''))
        ctype = (row.get('type', '') or '').strip().upper()
        field = (row.get('field', '') or '').strip()
        corrected = (row.get('corrected', '') or '').strip()
        original = (row.get('original', '') or '').strip()

        if not sid or not field:
            continue

        if sid not in effective:
            effective[sid] = {}

        if ctype == 'REVERT':
            effective[sid].pop(field, None)
        else:
            effective[sid][field] = {'corrected': corrected, 'original': original}

    # Flatten to rows the rest of the pipeline expects
    result_rows = []
    for sid, fields in effective.items():
        for field, vals in fields.items():
            corrected = vals['corrected']
            original = vals['original']
            if not corrected:
                continue
            # Map field names to what merge_corrections expects
            if field == 'student_id':
                question = 'ID'
            else:
                question = field  # e.g. "Q15", "LastName", etc.
            result_rows.append({
                'student_id': sid,
                'question': question,
                'corrected_answer': corrected,
                'original_answer': original,
            })

    if not result_rows:
        return pd.DataFrame(columns=['student_id', 'question', 'corrected_answer', 'original_answer'])

    df = pd.DataFrame(result_rows)
    print(f"[corrections] Loaded CSV with {len(df)} effective corrections", file=sys.stderr)
    return df


def _load_corrections_xlsx(corrections_xlsx: str) -> pd.DataFrame:
    """
    Load corrections from a filled-in flagged.xlsx file (legacy format).

    The file should have columns:
    - Student ID
    - Question
    - Corrected Answer (non-empty means a correction)

    Returns:
        DataFrame with columns: student_id, question, corrected_answer
        Only rows with non-empty Corrected Answer are returned.
    """
    import sys

    # Try 'Flagged Items' sheet first, then fall back to first sheet
    try:
        df = pd.read_excel(corrections_xlsx, sheet_name="Flagged Items")
    except ValueError:
        df = pd.read_excel(corrections_xlsx, sheet_name=0)
    print(f"[corrections] Loaded XLSX with columns: {list(df.columns)}", file=sys.stderr)

    # Normalize column names to handle variations
    col_map = {}
    for col in df.columns:
        col_lower = col.lower().replace(' ', '_').strip()
        if 'student' in col_lower and 'id' in col_lower:
            col_map[col] = 'student_id'
        elif col_lower == 'question':
            col_map[col] = 'question'
        elif 'corrected' in col_lower and 'answer' in col_lower:
            col_map[col] = 'corrected_answer'

    df = df.rename(columns=col_map)
    print(f"[corrections] After rename, columns: {list(df.columns)}", file=sys.stderr)

    # Filter to only rows with corrections
    if 'corrected_answer' not in df.columns:
        print("[corrections] Warning: No 'corrected_answer' column found!", file=sys.stderr)
        return pd.DataFrame(columns=['student_id', 'question', 'corrected_answer'])

    # Clean student_id and corrected_answer columns (handle Excel numeric conversion)
    if 'student_id' in df.columns:
        df['student_id'] = df['student_id'].apply(_clean_id_string)
    df['corrected_answer'] = df['corrected_answer'].apply(_clean_id_string)

    # Filter out rows without corrections
    mask = df['corrected_answer'] != ''
    corrections = df[mask]

    print(f"[corrections] Found {len(corrections)} rows with corrections", file=sys.stderr)
    if not corrections.empty:
        print(f"[corrections] First correction: {corrections.iloc[0].to_dict()}", file=sys.stderr)

    if corrections.empty:
        return pd.DataFrame(columns=['student_id', 'question', 'corrected_answer'])

    result = corrections[['student_id', 'question', 'corrected_answer']].copy()
    # Legacy XLSX format doesn't store original values — add empty column
    if 'original_answer' not in result.columns:
        result['original_answer'] = ''
    return result


def load_corrections(corrections_path: str) -> pd.DataFrame:
    """
    Load corrections from either CSV (new format) or XLSX (legacy format).

    Dispatches based on file extension:
    - ``.csv`` — new append-only format with ``# applies to:`` header
    - ``.xlsx`` — legacy flagged-items Excel format

    Returns:
        DataFrame with columns: student_id, question, corrected_answer
    """
    import sys

    path_lower = str(corrections_path).lower()
    if path_lower.endswith('.csv'):
        print(f"[corrections] Loading CSV corrections: {corrections_path}", file=sys.stderr)
        return _load_corrections_csv(corrections_path)
    else:
        print(f"[corrections] Loading XLSX corrections: {corrections_path}", file=sys.stderr)
        return _load_corrections_xlsx(corrections_path)


def _question_to_col(question_val, item_cols: List[str]) -> Optional[str]:
    """
    Map a question identifier (from the flagged XLSX) to a CSV column name.

    Handles: bare integer 22 -> 'Q22', string 'Q22' -> 'Q22', string '22' -> 'Q22'.
    """
    q_str = str(question_val).strip()

    # Direct match (e.g., 'Q22' in item_cols)
    if q_str in item_cols:
        return q_str
    if q_str.upper() in [c.upper() for c in item_cols]:
        for c in item_cols:
            if c.upper() == q_str.upper():
                return c

    # Bare number -> Q-prefixed (e.g., 22 -> 'Q22')
    q_str_digits = q_str.lstrip('Qq')
    if q_str_digits.isdigit():
        candidate = f"Q{int(q_str_digits)}"
        if candidate in item_cols:
            return candidate

    return None


def merge_corrections(
    df: pd.DataFrame,
    corrections: pd.DataFrame,
    item_cols: List[str],
    key_row_idx: int,
    roster: Optional[pd.DataFrame] = None,
) -> Tuple[pd.DataFrame, int]:
    """
    Apply corrections to the results DataFrame.

    For each correction, updates the corresponding question answer
    in the student's row and recalculates the scoring columns.

    Special handling for ID corrections (Question="ID"):
    - Updates the StudentID column to the corrected value
    - If roster provided, also updates name columns from roster

    Args:
        df: Results DataFrame from scored CSV
        corrections: DataFrame with columns: student_id, question, corrected_answer
        item_cols: List of question column names (Q1, Q2, etc.)
        key_row_idx: Index of the KEY row
        roster: Optional roster DataFrame for ID corrections

    Returns:
        Tuple of (modified DataFrame, number of corrections applied)
    """
    import sys

    if corrections.empty:
        return df, 0

    df = df.copy()
    corrections_applied = 0

    # Get the key values for recalculating scores
    key_row = df.iloc[key_row_idx]

    # Find VALUE rows for point values per question.
    # Build a lookup: version → {col: points}
    _value_mask = df.apply(
        lambda row: any(str(cell).strip().upper() == 'VALUE' for cell in row), axis=1
    )
    _version_col = _find_col(df, ['version', 'Version'])
    _version_point_values: Dict[str, Dict[str, float]] = {}
    for vidx in df[_value_mask].index:
        vrow = df.iloc[vidx]
        ver = str(vrow.get(_version_col, '')).strip() if _version_col else ''
        pv: Dict[str, float] = {}
        for col in item_cols:
            try:
                pv[col] = float(vrow.get(col, 1))
            except (ValueError, TypeError):
                pv[col] = 1.0
        _version_point_values[ver] = pv

    # For single-version CSVs, also try using the VALUE row right after key_row_idx
    if not _version_point_values and key_row_idx + 1 < len(df):
        maybe_val = df.iloc[key_row_idx + 1]
        if any(str(cell).strip().upper() == 'VALUE' for cell in maybe_val):
            pv = {}
            for col in item_cols:
                try:
                    pv[col] = float(maybe_val.get(col, 1))
                except (ValueError, TypeError):
                    pv[col] = 1.0
            _version_point_values[''] = pv

    # Find the student ID column
    id_col = _find_col(df, ['studentid', 'StudentID', 'Student_ID', 'student_id', 'ID', 'id'])

    # Page column — used as a fallback key for Simple Grade mode where
    # corrections are stored with the page number in the student_id field.
    page_col = _find_col(df, ['Page', 'page'])

    if id_col is None and page_col is None:
        print(f"[corrections] Warning: No student ID or Page column found in {list(df.columns)}", file=sys.stderr)
        return df, 0

    # Find name columns for ID corrections
    firstname_col = _find_col(df, ['firstname', 'FirstName', 'First_Name', 'first_name', 'First'])
    lastname_col = _find_col(df, ['lastname', 'LastName', 'Last_Name', 'last_name', 'Last'])

    # Pre-normalize all IDs in the DataFrame for matching
    df_ids_normalized = df[id_col].apply(_normalize_id) if id_col else pd.Series("", index=df.index)

    for _, correction in corrections.iterrows():
        student_id_normalized = _normalize_id(correction['student_id'])
        question_str = str(correction['question']).strip().upper()
        new_value = str(correction['corrected_answer']).strip()

        # Special handling for ID corrections (orphan scans)
        if question_str == 'ID':
            # This is an ID correction - update the StudentID column
            new_id = new_value.upper() if new_value else ''

            # Find the student row by their ORIGINAL (wrong) ID
            student_mask = df_ids_normalized == student_id_normalized
            _value_indices = set(df[_value_mask].index.tolist())
            student_indices = [i for i in df[student_mask].index.tolist()
                              if i != key_row_idx and i not in _value_indices]

            # Fallback: try matching by Page column (Simple Grade mode
            # stores the page number in the correction's student_id field).
            if not student_indices and page_col:
                page_mask = df[page_col].astype(str).str.strip() == student_id_normalized
                student_indices = [i for i in df[page_mask].index.tolist()
                                   if i != key_row_idx and i not in _value_indices]

            if not student_indices:
                sample_ids = df_ids_normalized[df_ids_normalized != 'KEY'].head(5).tolist()
                print(f"[corrections] Warning: Could not find student with ID '{student_id_normalized}' "
                      f"for ID correction (sample IDs: {sample_ids})", file=sys.stderr)
                continue

            for idx in student_indices:
                old_id = str(df.at[idx, id_col])
                df.at[idx, id_col] = new_id
                corrections_applied += 1
                print(f"[corrections] ID Correction: '{old_id}' -> '{new_id}'", file=sys.stderr)

                # Also update name from roster if available
                if roster is not None and new_id:
                    roster_match = roster[roster['StudentID'].apply(_normalize_id) == _normalize_id(new_id)]
                    if not roster_match.empty:
                        roster_row = roster_match.iloc[0]
                        if firstname_col and 'FirstName' in roster_row:
                            old_first = df.at[idx, firstname_col]
                            df.at[idx, firstname_col] = roster_row['FirstName']
                            print(f"[corrections]   FirstName: '{old_first}' -> '{roster_row['FirstName']}'", file=sys.stderr)
                        if lastname_col and 'LastName' in roster_row:
                            old_last = df.at[idx, lastname_col]
                            df.at[idx, lastname_col] = roster_row['LastName']
                            print(f"[corrections]   LastName: '{old_last}' -> '{roster_row['LastName']}'", file=sys.stderr)

            # Update normalized IDs cache since we changed an ID
            df_ids_normalized = df[id_col].apply(_normalize_id)
            continue

        # Normal answer correction
        new_answer = new_value.upper()

        # Map question to column name
        q_col = _question_to_col(correction['question'], item_cols)
        if q_col is None:
            print(f"[corrections] Warning: Could not find column for question '{correction['question']}' "
                  f"(available: {item_cols[:5]}...)", file=sys.stderr)
            continue

        # Find matching student rows (excluding KEY and VALUE rows)
        student_mask = df_ids_normalized == student_id_normalized
        _value_indices = set(df[_value_mask].index.tolist())
        student_indices = [i for i in df[student_mask].index.tolist()
                          if i != key_row_idx and i not in _value_indices]

        # Fallback: try matching by Page column (Simple Grade mode
        # stores the page number in the correction's student_id field).
        if not student_indices and page_col:
            page_mask = df[page_col].astype(str).str.strip() == student_id_normalized
            student_indices = [i for i in df[page_mask].index.tolist()
                               if i != key_row_idx and i not in _value_indices]

        if not student_indices:
            sample_ids = df_ids_normalized[df_ids_normalized != 'KEY'].head(5).tolist()
            print(f"[corrections] Warning: Could not find student '{student_id_normalized}' "
                  f"(sample IDs: {sample_ids})", file=sys.stderr)
            continue

        # Apply correction
        for idx in student_indices:
            old_answer = str(df.at[idx, q_col]).upper().strip()
            df.at[idx, q_col] = new_answer
            corrections_applied += 1
            print(f"[corrections] Applied: Student {student_id_normalized}, "
                  f"{q_col}: '{old_answer}' -> '{new_answer}'", file=sys.stderr)

            # Look up point values for this student's version
            student_ver = ''
            if _version_col:
                student_ver = str(df.at[idx, _version_col]).strip().rstrip("*")
            pv = _version_point_values.get(student_ver) or _version_point_values.get('') or None
            _recalculate_student_scores(df, idx, item_cols, key_row, point_values=pv)

    return df, corrections_applied


def _answer_matches_key(answer: str, key_answer: str) -> bool:
    """Check if *answer* matches *key_answer*, handling compound keys.

    Compound key formats: AND ``B&E``, OR ``B^D``, partial ``A@B`` / ``A~B``.
    Student multi-mark answers use commas: ``B,E``.
    """
    import re as _re
    a = answer.strip().upper()
    k = key_answer.strip().upper()
    if not a or not k or a in ("NAN", "NONE") or k in ("NAN", "NONE"):
        return False

    # Normalise student answer to a set
    student_set = frozenset(p.strip() for p in a.split(",") if p.strip())

    if "&" in k:
        key_set = frozenset(p.strip() for p in k.split("&") if p.strip())
        return student_set == key_set
    if _re.search(r"[@~]", k):
        sep = _re.search(r"[@~]", k).group()
        key_set = frozenset(p.split(":")[0].strip() for p in k.split(sep) if p.strip())
        return student_set == key_set
    if "^" in k:
        accepted = frozenset(p.strip() for p in k.split("^") if p.strip())
        return len(student_set) == 1 and bool(student_set & accepted)

    # Simple single-answer key
    return len(student_set) == 1 and a == k


def _recalculate_student_scores(
    df: pd.DataFrame,
    student_idx: int,
    item_cols: List[str],
    key_row: pd.Series,
    point_values: Optional[Dict[str, float]] = None,
):
    """
    Recalculate correct/incorrect/blank/multi/percent/score for a student
    after corrections.

    Args:
        point_values: Optional dict mapping column name → max points for that
            question (from the VALUE row in the CSV).  When provided, *score*
            is the sum of earned points (correct × per-question value), and
            *percent* is score / max_points × 100.  When ``None``, every
            question is worth 1 point (legacy behaviour).
    """
    correct = 0
    incorrect = 0
    blank = 0
    multi = 0
    score = 0.0
    max_points = 0.0

    for col in item_cols:
        answer = str(df.at[student_idx, col]).upper().strip()
        key_answer = str(key_row[col]).upper().strip()
        pts = (point_values or {}).get(col, 1.0)
        max_points += pts

        if not answer or answer in ('', 'NAN', 'NONE', 'BLANK', '-'):
            blank += 1
        elif _answer_matches_key(answer, key_answer):
            correct += 1
            score += pts
        elif ',' in answer:
            # Multi-mark that did NOT match a compound key → unexpected multi
            multi += 1
        else:
            incorrect += 1

    percent = (score / max_points * 100) if max_points > 0 else 0

    # Update the score columns (handle all case variants from CSV normalization)
    score_val = score if score != int(score) else int(score)
    for name, value in [('score', score_val), ('correct', correct), ('incorrect', incorrect),
                        ('blank', blank), ('multi', multi), ('percent', round(percent, 2))]:
        col_name = _find_col(df, [name, name.capitalize(), name.upper()])
        if col_name is not None:
            df.at[student_idx, col_name] = value


def apply_corrections_to_csv(
    input_csv: str,
    corrections_xlsx: str,
    output_csv: str,
    roster_csv: Optional[str] = None,
) -> int:
    """
    Apply teacher corrections to a scored CSV and write a new corrected CSV.

    Reads the scored CSV, loads corrections from a corrections file
    (.csv or .xlsx), updates the answer columns, recalculates scores,
    and writes a new CSV.

    Args:
        input_csv: Path to original scored CSV
        corrections_xlsx: Path to corrections file (.csv or .xlsx)
        output_csv: Path to write the corrected CSV
        roster_csv: Optional roster CSV for ID corrections (to look up names)

    Returns:
        Number of corrections applied
    """
    import sys

    # Load the CSV
    df = _load_score_csv_robust(input_csv)

    # Detect item columns and key row
    item_cols = detect_item_columns(df, r"Q\d+")
    if not item_cols:
        raise ValueError(f"No item columns (Q1, Q2...) found in CSV. Columns: {list(df.columns)}")

    key_row_idx = detect_key_row_index(df, item_cols, key_label="KEY")

    # Load roster if provided (for ID corrections)
    roster = None
    if roster_csv:
        try:
            roster = load_roster(roster_csv)
            print(f"[corrections] Loaded roster with {len(roster)} students for ID corrections", file=sys.stderr)
        except Exception as e:
            print(f"[corrections] Warning: Could not load roster: {e}", file=sys.stderr)

    # Load and apply corrections
    corrections = load_corrections(corrections_xlsx)
    if corrections.empty:
        print("[corrections] No corrections found in XLSX", file=sys.stderr)
        # Still write the output (a clean copy)
        df.to_csv(output_csv, index=False)
        return 0

    df, corrections_applied = merge_corrections(df, corrections, item_cols, key_row_idx, roster=roster)

    # Write corrected CSV
    df.to_csv(output_csv, index=False)
    print(f"[corrections] Wrote corrected CSV with {corrections_applied} corrections to {output_csv}", file=sys.stderr)

    return corrections_applied


# ==================== ROSTER MATCHING ====================

def load_roster(roster_path: str) -> pd.DataFrame:
    """
    Load and normalize a class roster CSV.

    Expected columns (case-insensitive, auto-detected):
    - StudentID / ID / Student_ID
    - LastName / Last / Surname
    - FirstName / First (optional)

    Returns DataFrame with standardized columns: StudentID, LastName, FirstName
    """
    df = pd.read_csv(roster_path)

    # Normalize column names
    col_map = {}
    for col in df.columns:
        col_lower = col.lower().strip()
        if col_lower in ('studentid', 'id', 'student_id', 'sid'):
            col_map[col] = 'StudentID'
        elif col_lower in ('lastname', 'last', 'surname', 'last_name'):
            col_map[col] = 'LastName'
        elif col_lower in ('firstname', 'first', 'first_name'):
            col_map[col] = 'FirstName'

    if 'StudentID' not in col_map.values():
        raise ValueError(
            f"Roster CSV must have a student ID column. "
            f"Expected: StudentID/ID/Student_ID. Found: {list(df.columns)}"
        )

    if 'LastName' not in col_map.values():
        raise ValueError(
            f"Roster CSV must have a last name column. "
            f"Expected: LastName/Last/Surname. Found: {list(df.columns)}"
        )

    df = df.rename(columns=col_map)

    # Fill missing FirstName with empty string
    if 'FirstName' not in df.columns:
        df['FirstName'] = ''

    # Convert StudentID to string, strip whitespace, and remove float artifacts
    # (pandas reads numeric IDs as float when NaNs are present, producing "1234.0")
    df['StudentID'] = df['StudentID'].apply(_normalize_id)
    df['LastName'] = df['LastName'].astype(str).str.strip()
    df['FirstName'] = df['FirstName'].astype(str).str.strip()

    return df[['StudentID', 'LastName', 'FirstName']]


def fuzzy_match_student(
    scanned_id: str,
    scanned_last: str,
    scanned_first: str,
    roster: pd.DataFrame,
    id_threshold: float = 85.0,
    name_threshold: float = 85.0,
) -> Tuple[Optional[str], float, str]:
    """
    Attempt to match a scanned student to the roster using fuzzy matching.

    Returns:
        (matched_roster_id, confidence, match_type)

        match_type can be:
        - "exact": Exact StudentID match
        - "high_confidence": High ID similarity or ID + name match
        - "probable": Moderate confidence match
        - "no_match": No good match found
    """
    if roster.empty:
        return None, 0.0, "no_match"

    # Clean inputs
    scanned_id = str(scanned_id).strip()
    scanned_last = str(scanned_last).strip().upper()
    scanned_first = str(scanned_first).strip().upper()

    best_match = None
    best_score = 0.0
    match_type = "no_match"

    for _, row in roster.iterrows():
        roster_id = str(row['StudentID']).strip()
        roster_last = str(row['LastName']).strip().upper()
        roster_first = str(row['FirstName']).strip().upper()

        # Exact ID match
        if scanned_id == roster_id:
            return roster_id, 100.0, "exact"

        # Fuzzy ID match
        id_score = fuzz.ratio(scanned_id, roster_id)

        # Name matching
        last_score = fuzz.ratio(scanned_last, roster_last) if scanned_last else 0
        first_score = fuzz.ratio(scanned_first, roster_first) if scanned_first and roster_first else 0

        # Combined scoring strategies
        # Strategy 1: Very high ID match (typo in one digit)
        if id_score >= 95:
            confidence = id_score
            if confidence > best_score:
                best_match = roster_id
                best_score = confidence
                match_type = "high_confidence"

        # Strategy 2: Good ID match + exact last name
        elif id_score >= id_threshold and last_score == 100:
            confidence = (id_score + last_score) / 2
            if confidence > best_score:
                best_match = roster_id
                best_score = confidence
                match_type = "high_confidence"

        # Strategy 3: Perfect name match (both first and last)
        elif last_score == 100 and first_score == 100 and scanned_first:
            confidence = 100.0
            if confidence > best_score:
                best_match = roster_id
                best_score = confidence
                match_type = "probable"

        # Strategy 4: Good overall match
        elif id_score >= id_threshold or (last_score >= name_threshold and id_score >= 70):
            confidence = max(id_score, (id_score + last_score) / 2)
            if confidence > best_score:
                best_match = roster_id
                best_score = confidence
                match_type = "probable"

    return best_match, best_score, match_type


def match_students_to_roster(
    students_df: pd.DataFrame,
    roster_df: pd.DataFrame,
) -> Tuple[pd.DataFrame, List[Dict], List[Dict]]:
    """
    Match scanned students to roster using **exact ID matching only**.

    Fuzzy matching is used solely to populate hint columns (RosterID,
    MatchConfidence, MatchType) that feed the orphan-suggestions UI in the
    Review panel.  It never affects the matched / orphan / absent
    classification — only an exact StudentID match counts.

    The teacher must explicitly accept a correction in the Review panel
    before an orphan is treated as matched.

    Returns:
        (students_df_with_matches, orphan_scans, absent_students)

        students_df_with_matches: Original DataFrame with added columns:
            - RosterID: Fuzzy-matched roster ID hint (or None)
            - MatchConfidence: 0-100
            - MatchType: exact/high_confidence/probable/no_match

        orphan_scans: List of dicts for students whose ID is not in roster
        absent_students: List of dicts for roster students with no exact scan match
    """
    # Add hint columns (informational only — do not drive matching logic)
    students_df['RosterID'] = None
    students_df['MatchConfidence'] = 0.0
    students_df['MatchType'] = 'no_match'

    # Build a set of normalised roster IDs for O(1) exact-match lookup
    roster_id_set = {_normalize_id(row['StudentID']) for _, row in roster_df.iterrows()}

    # IDs with an exact match — used to determine absent students
    matched_roster_ids = set()

    # Resolve column names (scored CSV may use lowercase 'studentid', etc.)
    _sid_col = _find_col(students_df, ['StudentID', 'studentid', 'Student_ID', 'student_id', 'ID', 'id'])
    _last_col = _find_col(students_df, ['LastName', 'lastname', 'Last', 'last', 'Surname', 'last_name'])
    _first_col = _find_col(students_df, ['FirstName', 'firstname', 'First', 'first', 'first_name'])

    # Build absent_roster dict for suggest_matches() — initially all roster
    # entries; we'll remove exact matches as we find them.
    _absent_for_suggest: Dict[str, Dict[str, str]] = {}
    for _, rrow in roster_df.iterrows():
        rid = _normalize_id(rrow['StudentID'])
        _absent_for_suggest[rid] = {
            'full_name': f"{rrow.get('LastName', '')}, {rrow.get('FirstName', '')}".strip(", "),
        }

    # Import the unified suggest_matches from score_core
    from ..score_core import suggest_matches as _suggest_matches

    for idx, row in students_df.iterrows():
        scanned_id = _normalize_id(row.get(_sid_col, '')) if _sid_col else ''
        scanned_last = str(row.get(_last_col, '')).strip() if _last_col else ''
        scanned_first = str(row.get(_first_col, '')).strip() if _first_col else ''

        # Exact match — the only thing that counts
        if scanned_id and scanned_id in roster_id_set:
            students_df.at[idx, 'RosterID'] = scanned_id
            students_df.at[idx, 'MatchConfidence'] = 100.0
            students_df.at[idx, 'MatchType'] = 'exact'
            matched_roster_ids.add(scanned_id)
            # Remove from absent pool so it isn't suggested for other orphans
            _absent_for_suggest.pop(scanned_id, None)
            continue

        # No exact match → orphan.  Run unified fuzzy match for hint info only.
        if scanned_id or scanned_last:
            orphan_name = f"{scanned_last}, {scanned_first}".strip(", ")
            matches = _suggest_matches(
                orphan_id=scanned_id,
                orphan_name=orphan_name,
                absent_roster=_absent_for_suggest,
                max_suggestions=1,
            )
            if matches:
                top = matches[0]
                students_df.at[idx, 'RosterID'] = top['student_id']
                students_df.at[idx, 'MatchConfidence'] = float(top['score'])
                # Override type — regardless of fuzzy confidence this is
                # still an orphan; the hint is stored but does not promote
                # the scan to "matched".
                students_df.at[idx, 'MatchType'] = 'probable'

    # Orphan scans — every non-exact match
    orphan_scans = []
    for idx, row in students_df.iterrows():
        if row['MatchType'] != 'exact':
            orphan_scans.append({
                'ScannedID': _normalize_id(row.get(_sid_col, '')) if _sid_col else '',
                'LastName': str(row.get(_last_col, '')).strip() if _last_col else '',
                'FirstName': str(row.get(_first_col, '')).strip() if _first_col else '',
                'MatchType': row['MatchType'],
                'PossibleMatch': row['RosterID'] if row['MatchType'] == 'probable' else None,
                'Confidence': row['MatchConfidence'],
            })

    # Absent students — roster entries with no exact scan match
    absent_students = []
    for _, row in roster_df.iterrows():
        roster_id = _normalize_id(row['StudentID'])
        if roster_id not in matched_roster_ids:
            absent_students.append({
                'StudentID': roster_id,
                'LastName': row['LastName'],
                'FirstName': row['FirstName'],
            })

    return students_df, orphan_scans, absent_students


# ==================== EXCEL FORMATTING ====================

# Color scheme
COLOR_HEADER = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
COLOR_KEY_ROW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
COLOR_GOOD = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
COLOR_WARNING = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
COLOR_PROBLEM = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
COLOR_ORPHAN = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
COLOR_BLANK = PatternFill(start_color="F056E3", end_color="F056E3", fill_type="solid")  # Pink-purple for blank answers
COLOR_MULTI = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")  # Orange for multi-answer

FONT_HEADER = Font(bold=True, color="FFFFFF")
FONT_BOLD = Font(bold=True)
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def format_header_row(ws, row_num: int):
    """Apply header formatting to a row."""
    for cell in ws[row_num]:
        cell.fill = COLOR_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN


def format_key_row(ws, row_num: int):
    """Apply KEY row formatting."""
    for cell in ws[row_num]:
        cell.fill = COLOR_KEY_ROW
        cell.font = FONT_BOLD
        cell.border = BORDER_THIN


def auto_size_columns(ws):
    """Auto-size all columns based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


# ==================== CSV LOADING ====================

def _load_score_csv_robust(input_csv: str) -> pd.DataFrame:
    """
    Load CSV from score command, handling messy format with include_stats.

    The score command with include_stats can create CSVs with:
    - Section headers like "=== VERSION A (112 students) ==="
    - Multiple header rows (one per version)
    - Stats rows at the bottom

    This function:
    1. Finds the first valid header row
    2. Reads only the student data rows (skipping section headers and stats)
    3. Returns a clean DataFrame
    """
    import csv

    # Read the raw CSV to find structure
    with open(input_csv, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    # Find the first valid header row (has question columns like Q1, Q2, etc.)
    header_idx = None
    header = None

    for idx, line in enumerate(lines):
        row = next(csv.reader([line]))
        # Valid header should have StudentID and Q1
        if 'StudentID' in row or 'StudentID' in str(row):
            # Check if this looks like a real header with questions
            if any(col.startswith('Q') and col[1:].isdigit() for col in row if isinstance(col, str)):
                header_idx = idx
                header = row
                break

    if header_idx is None:
        # Fallback: just try reading normally
        return pd.read_csv(input_csv)

    # Collect all data rows (skip section headers and stats rows)
    data_rows = []
    reader = csv.reader(lines[header_idx + 1:])

    for row in reader:
        # Skip empty rows
        if not row or all(cell.strip() == '' for cell in row):
            continue

        # Skip section headers (=== VERSION A === etc.)
        if row and row[0].strip().startswith('==='):
            continue

        # Skip stats rows (start with specific labels)
        if row and row[0].strip() in ['PCT_CORRECT', 'POINT_BISERIAL', 'N_STUDENTS',
                                        'MEAN_SCORE', 'MEAN_PERCENT', 'STD_DEV',
                                        'HIGH_SCORE', 'LOW_SCORE', 'KR20_OVERALL'] or \
           (row and 'PCT_CORRECT' in row[0]) or \
           (row and 'POINT_BISERIAL' in row[0]) or \
           (row and 'KR20_VERSION' in row[0]) or \
           (row and '--- ITEM STATISTICS' in row[0]):
            continue

        # Skip duplicate header rows (multi-version CSVs repeat the header)
        if row == header:
            continue
        # Also catch headers that match loosely (e.g., same first few columns)
        if len(row) >= 3 and row[0].strip() == header[0].strip() and row[1].strip() == header[1].strip() and row[2].strip() == header[2].strip():
            continue

        # This looks like a valid data row
        # Ensure row has same length as header
        if len(row) < len(header):
            row.extend([''] * (len(header) - len(row)))
        elif len(row) > len(header):
            row = row[:len(header)]

        data_rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(data_rows, columns=header)

    # Normalize column names to lowercase for consistency
    # This handles both 'Correct' and 'correct', 'Version' and 'version', etc.
    column_mapping = {}
    for col in df.columns:
        col_lower = col.lower().strip()
        # Map common variations to standard names
        if col_lower in ('score', 'correct', 'incorrect', 'blank', 'multi', 'percent',
                         'version', 'studentid', 'lastname', 'firstname',
                         'page', 'page_index'):
            column_mapping[col] = col_lower
        # Keep question columns as-is (Q1, Q2, etc.)
        elif col.startswith('Q') and col[1:].isdigit():
            column_mapping[col] = col
        else:
            column_mapping[col] = col

    df = df.rename(columns=column_mapping)

    # Convert score columns from strings to proper numeric types so Excel
    # formats them as numbers instead of left-aligned text.
    for col in ('score', 'correct', 'incorrect', 'blank', 'multi'):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    if 'percent' in df.columns:
        df['percent'] = pd.to_numeric(df['percent'], errors='coerce')

    return df


# ==================== REPORT GENERATION ====================

def generate_report(
    input_csv: str,
    out_xlsx: str,
    roster_csv: Optional[str] = None,
    item_pattern: str = r"Q\d+",
    project_name: Optional[str] = None,
    run_label: Optional[str] = None,
    corrections_applied: int = 0,
    corrections_xlsx: Optional[str] = None,
    scoring_params: Optional[Dict] = None,
    simple: bool = False,
):
    r"""
    Generate comprehensive Excel report from scored CSV.

    Args:
        input_csv: Path to scored CSV from markshark score
        out_xlsx: Path to output Excel file
        roster_csv: Optional path to class roster CSV
        item_pattern: Regex pattern for item columns (default: Q\d+)
        project_name: Optional project name for report header
        run_label: Optional run label (e.g., "2025-01-21_final")
        corrections_applied: Number of corrections applied (for display on Summary tab)
        corrections_xlsx: Optional path to the corrections XLSX (for listing details on Summary tab)
        simple: When True, produce a streamlined report with only a Class
            Scores sheet and Answer Key sheet (no per-version item analysis,
            no roster matching, no detailed statistics).  Designed for
            small-class "Simple Grade" workflows.
    """
    # Auto-load scoring parameters from companion JSON if not explicitly given
    if scoring_params is None:
        import json as _json
        params_path = os.path.splitext(input_csv)[0] + "_params.json"
        if os.path.isfile(params_path):
            try:
                with open(params_path, "r", encoding="utf-8") as pf:
                    scoring_params = _json.load(pf)
            except Exception:
                scoring_params = None

    # Load scored results - handle messy CSV from score with include_stats
    df = _load_score_csv_robust(input_csv)

    # Detect item columns and key row
    item_cols = detect_item_columns(df, item_pattern)
    if not item_cols:
        raise ValueError(
            f"No item columns found matching pattern '{item_pattern}'. "
            f"Available columns: {list(df.columns)}"
        )

    k = len(item_cols)

    # Apply corrections to the data BEFORE computing stats / building tabs
    if corrections_xlsx:
        corrections_df = load_corrections(corrections_xlsx)
        if not corrections_df.empty:
            key_row_idx = detect_key_row_index(df, item_cols, key_label="KEY")
            # roster_df isn't loaded yet; load a lightweight copy for ID corrections
            roster_for_merge = None
            if roster_csv:
                try:
                    roster_for_merge = load_roster(roster_csv)
                except Exception:
                    pass
            df, corrections_applied = merge_corrections(
                df, corrections_df, item_cols, key_row_idx, roster=roster_for_merge,
            )

    # ── Simple Grade shortcut ──────────────────────────────────────────
    # Skip roster matching, item analysis, and per-version stats.
    # Produce only a Class Scores sheet and an Answer Key sheet.
    if simple:
        # We still need version info for the answer-key tab, but we can
        # compute it cheaply without the full per-version stats pipeline.
        version_col = _find_col(df, ['Version', 'version'])
        if version_col:
            all_v = df[version_col].dropna().astype(str).str.strip().unique()
            versions = sorted({
                v.rstrip("*") for v in all_v
                if v.rstrip("*") and len(v.rstrip("*")) <= 2
                and v.rstrip("*").upper() not in ('VERSION', 'KEY')
            })
        else:
            versions = ['A']

        # Build minimal version_stats with just the key answers (for the
        # Answer Key tab).  No difficulty, no point-biserial, no KR-20.
        version_stats: Dict[str, dict] = {}
        for ver in versions:
            if version_col:
                ver_mask = df[version_col].astype(str).str.strip() == ver
                df_ver = df[ver_mask]
            else:
                df_ver = df
            key_idx = detect_key_row_index(df_ver, item_cols, key_label="KEY")
            if key_idx is not None:
                key_row = df_ver.iloc[key_idx]
                key_series = pd.Series({
                    col: str(key_row[col]).strip() for col in item_cols
                })
            else:
                key_series = pd.Series({col: "" for col in item_cols})
            version_stats[ver] = {'key_series': key_series}

        wb = Workbook()
        wb.remove(wb.active)

        # ── Simple Summary sheet (lightweight header only) ──
        ws = wb.create_sheet("Summary")
        ws.cell(row=1, column=1, value="MarkShark Simple Grade Report")
        ws.cell(row=1, column=1).font = FONT_BOLD
        row_n = 3
        if project_name:
            ws.cell(row=row_n, column=1, value="Assessment:")
            ws.cell(row=row_n, column=2, value=project_name)
            row_n += 1
        from datetime import datetime as _dt
        ws.cell(row=row_n, column=1, value="Generated:")
        ws.cell(row=row_n, column=2, value=_dt.now().strftime("%Y-%m-%d %H:%M"))
        row_n += 1
        ws.cell(row=row_n, column=1, value="Total questions:")
        ws.cell(row=row_n, column=2, value=k)
        row_n += 1
        if corrections_applied:
            ws.cell(row=row_n, column=1, value="Corrections applied:")
            ws.cell(row=row_n, column=2, value=corrections_applied)
        auto_size_columns(ws)

        create_class_scores_tab(wb, df, item_cols, k)
        create_answer_key_tab(wb, item_cols, versions, version_stats)

        wb.save(out_xlsx)
        print(f"Simple Grade report generated: {out_xlsx}")
        return

    # Load roster if provided
    roster_df = None
    orphan_scans = []
    absent_students = []

    if roster_csv:
        roster_df = load_roster(roster_csv)
        # Remove all KEY/VALUE rows before matching (multi-version CSVs have one per version)
        key_mask = df.apply(
            lambda row: any(str(cell).strip().upper() in ('KEY', 'VALUE') for cell in row), axis=1
        )
        students_only = df[~key_mask].reset_index(drop=True)
        students_only, orphan_scans, absent_students = match_students_to_roster(
            students_only, roster_df
        )

    # Group by version (handle both 'Version' and 'version' column names)
    version_col = None
    if 'version' in df.columns:
        version_col = 'version'
    elif 'Version' in df.columns:
        version_col = 'Version'

    if version_col:
        # Get unique versions, filtering out invalid values.
        # Strip the auto-detect marker (*) so "B*" maps to base version "B".
        all_versions = df[version_col].dropna().astype(str).str.strip().unique()
        base_versions: set = set()
        for v in all_versions:
            v_clean = v.rstrip("*")
            if v_clean and len(v_clean) <= 2 and v_clean.upper() not in ('VERSION', 'KEY'):
                base_versions.add(v_clean)
        versions = sorted(base_versions)
    else:
        versions = ['A']  # Default single version

    # Compute per-version statistics
    # Each version has its own answer key, so correctness must be computed per-version.
    # Pooled (amalgamated) stats are then derived by combining per-version results.
    version_stats = {}
    all_items_num = []        # Collect 0/1 correctness matrices (for KR-20, point-biserial)
    all_total_scores = []     # Collect correctness sums (for KR-20, point-biserial)
    all_weighted_scores = []  # Collect point-weighted scores from CSV (for mean/std display)
    total_n_students = 0

    for version in versions:
        # Filter to only students who took this version.
        # Match both "B" and "B*" (auto-detected) against base version "B".
        if version_col:
            version_mask = df[version_col].astype(str).str.strip().str.rstrip("*") == str(version).strip()
            df_version = df[version_mask].copy()
        else:
            df_version = df.copy()

        # Extract VALUE row data before dropping (for total points calculation)
        value_mask_v = df_version.apply(
            lambda row: any(str(cell).strip().upper() == 'VALUE' for cell in row), axis=1
        )
        total_pts = float(k)  # default: k questions × 1 point each
        if value_mask_v.any():
            val_row = df_version[value_mask_v].iloc[0]
            total_pts = 0.0
            for col in item_cols:
                try:
                    total_pts += float(val_row.get(col, 1))
                except (ValueError, TypeError):
                    total_pts += 1.0

        # Drop VALUE rows before stats (prepare_correctness_matrix only drops KEY)
        df_version = df_version[~value_mask_v].reset_index(drop=True)

        # Find KEY row for this version
        key_row_idx_version = detect_key_row_index(df_version, item_cols, key_label="KEY")

        # Prepare correctness matrix for this version only
        items_num_v, total_scores_v, students_df_v, key_series_v = prepare_correctness_matrix(
            df_version, item_cols, key_row_idx_version, answers_mode="auto"
        )

        n_students_v = len(students_df_v)
        total_n_students += n_students_v

        # Compute difficulty (% correct) for this version
        difficulty_v = items_num_v.mean(axis=0)

        # Compute point-biserial for this version
        pb_vals_v = {}
        for col in item_cols:
            item_series = items_num_v[col]
            total_minus = total_scores_v - item_series.fillna(0)
            pb_vals_v[col] = point_biserial(item_series, total_minus)

        # Version-level exam stats
        # Use the point-weighted Score column from the CSV for mean/std
        # (total_scores_v is the 0/1 correctness sum, which only equals the
        # real score when all questions are worth 1 point).
        score_col = _find_col(students_df_v, ['Score', 'score', 'SCORE'])
        if score_col is not None:
            weighted_scores_v = pd.to_numeric(students_df_v[score_col], errors='coerce').fillna(0)
        else:
            # Fallback: no Score column → use correctness sum (all questions worth 1)
            weighted_scores_v = total_scores_v

        mean_v = float(weighted_scores_v.mean()) if n_students_v > 0 else 0.0
        std_v = float(weighted_scores_v.std(ddof=1)) if n_students_v > 1 else 0.0
        kr20_v = kr20(items_num_v, total_scores_v)
        kr21_v = kr21(items_num_v, total_scores_v)

        # Store version-specific stats
        version_stats[version] = {
            'difficulty': difficulty_v,
            'pb_vals': pb_vals_v,
            'key_series': key_series_v,
            'n_students': n_students_v,
            'mean': mean_v,
            'std': std_v,
            'kr20': kr20_v,
            'kr21': kr21_v,
            'total_points': total_pts,
        }

        # Accumulate for pooled stats
        # items_num / total_scores are 0/1 correctness (for KR-20, point-biserial)
        # weighted_scores are point-weighted (for mean/std display)
        all_items_num.append(items_num_v)
        all_total_scores.append(total_scores_v)
        all_weighted_scores.append(weighted_scores_v)

    # Compute pooled (amalgamated) stats from per-version correctness matrices.
    # Each student was scored against their own version's key, so these are correct.
    if all_items_num:
        pooled_items = pd.concat(all_items_num, ignore_index=True)
        pooled_totals = pd.concat(all_total_scores, ignore_index=True)
        pooled_weighted = pd.concat(all_weighted_scores, ignore_index=True)
    else:
        pooled_items = pd.DataFrame(columns=item_cols)
        pooled_totals = pd.Series(dtype=float)
        pooled_weighted = pd.Series(dtype=float)

    # Mean/std use point-weighted scores (from the Score column in the CSV).
    # KR-20/KR-21 use the 0/1 correctness sums (psychometric formulas need binary item data).
    mean_total = float(pooled_weighted.mean()) if total_n_students > 0 else 0.0
    std_total = float(pooled_weighted.std(ddof=1)) if total_n_students > 1 else 0.0
    kr20_val = kr20(pooled_items, pooled_totals)
    kr21_val = kr21(pooled_items, pooled_totals)

    # Load corrections detail for display on summary tab
    corrections_detail = None
    if corrections_xlsx:
        try:
            if str(corrections_xlsx).lower().endswith('.csv'):
                # New CSV format — load effective corrections for display
                corrections_detail = load_corrections(corrections_xlsx)
                # Rename columns for display
                corrections_detail = corrections_detail.rename(columns={
                    'student_id': 'Student ID',
                    'question': 'Question',
                    'corrected_answer': 'Corrected Answer',
                    'original_answer': 'Original Answer',
                })
                # Add student names from the results DataFrame
                id_col = _find_col(df, ['studentid', 'StudentID', 'Student_ID', 'student_id', 'ID', 'id'])
                last_col = _find_col(df, ['lastname', 'LastName', 'Last_Name', 'last_name', 'Last'])
                first_col = _find_col(df, ['firstname', 'FirstName', 'First_Name', 'first_name', 'First'])
                if id_col and (last_col or first_col):
                    names = []
                    for _, crow in corrections_detail.iterrows():
                        sid = _normalize_id(crow.get('Student ID', ''))
                        mask = df[id_col].apply(_normalize_id) == sid
                        matched = df[mask]
                        if not matched.empty:
                            row_data = matched.iloc[0]
                            last_name = str(row_data.get(last_col, '')).strip() if last_col else ''
                            first_name = str(row_data.get(first_col, '')).strip() if first_col else ''
                            name = f"{last_name}, {first_name}".strip(', ')
                            names.append(name)
                        else:
                            names.append('')
                    corrections_detail.insert(1, 'Student Name', names)
            else:
                try:
                    corrections_detail = pd.read_excel(corrections_xlsx, sheet_name="Flagged Items")
                except ValueError:
                    corrections_detail = pd.read_excel(corrections_xlsx, sheet_name=0)
        except Exception:
            corrections_detail = None

    # ---- Collect students with auto-detected versions (Version:blank) ----
    auto_detected_students: List[Dict] = []
    if version_col and len(versions) > 1:
        _id_col = _find_col(df, ['studentid', 'StudentID', 'Student_ID', 'student_id', 'ID', 'id'])
        _last_col = _find_col(df, ['lastname', 'LastName', 'Last_Name', 'last_name', 'Last'])
        _first_col = _find_col(df, ['firstname', 'FirstName', 'First_Name', 'first_name', 'First'])

        star_mask = df[version_col].astype(str).str.strip().str.endswith("*")
        non_student_mask_ad = df.apply(
            lambda r: any(str(c).strip().upper() in ('KEY', 'VALUE') for c in r), axis=1
        )
        key_only_mask_ad = df.apply(
            lambda r: any(str(c).strip().upper() == 'KEY' for c in r), axis=1
        )
        star_students = df[star_mask & ~non_student_mask_ad]

        # Build per-version KEY answer lookup
        version_keys: Dict[str, Dict[str, str]] = {}
        for ver in versions:
            ver_mask = df[version_col].astype(str).str.strip() == ver
            key_rows = df[ver_mask & key_only_mask_ad]
            if not key_rows.empty:
                kr = key_rows.iloc[0]
                version_keys[ver] = {
                    col: str(kr[col]).strip().upper()
                    for col in item_cols
                    if pd.notna(kr.get(col))
                }

        for _, srow in star_students.iterrows():
            auto_ver = str(srow[version_col]).strip().rstrip("*")
            scores_by_ver: Dict[str, int] = {}
            for ver, keys in version_keys.items():
                correct = sum(
                    1 for col in item_cols
                    if _answer_matches_key(
                        str(srow[col]).strip() if pd.notna(srow.get(col)) else "",
                        keys.get(col, ""),
                    )
                )
                scores_by_ver[ver] = correct

            auto_detected_students.append({
                'student_id': _normalize_id(srow.get(_id_col, '')) if _id_col else '',
                'last_name': str(srow.get(_last_col, '')).strip() if _last_col else '',
                'first_name': str(srow.get(_first_col, '')).strip() if _first_col else '',
                'auto_version': auto_ver,
                'scores': scores_by_ver,
                'total': len(item_cols),
            })

    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ========== SUMMARY TAB ==========
    create_summary_tab(
        wb, k, total_n_students, mean_total, std_total, kr20_val, kr21_val,
        versions, version_stats,
        orphan_scans, absent_students, project_name, run_label,
        corrections_applied, corrections_detail,
        df=df, item_cols=item_cols, input_csv_path=input_csv,
        scoring_params=scoring_params,
        auto_detected_students=auto_detected_students,
    )

    # ========== PER-VERSION TABS ==========
    for version in versions:
        # Get version-specific stats
        vstats = version_stats[version]
        create_version_tab(
            wb, df, None, version, item_cols, vstats['key_series'],
            vstats['difficulty'], vstats['pb_vals'], roster_df, orphan_scans if roster_csv else None
        )

    # ========== CLASS SCORES TAB ==========
    create_class_scores_tab(wb, df, item_cols, k)

    # ========== ANSWER KEY TAB ==========
    create_answer_key_tab(wb, item_cols, versions, version_stats)

    # Save workbook
    wb.save(out_xlsx)
    print(f"Excel report generated: {out_xlsx}")


def create_summary_tab(
    wb, k, n_students, mean_total, std_total, kr20_val, kr21_val,
    versions, version_stats,
    orphan_scans, absent_students, project_name=None, run_label=None,
    corrections_applied=0, corrections_detail=None,
    df=None, item_cols=None, input_csv_path=None,
    scoring_params=None,
    auto_detected_students=None,
):
    """Create summary tab with statistics, flagged items, and scoring parameters.

    Layout order:
        1. Title
        2. Project / Scores file / Generated
        3. Per-version & amalgamated statistics + reliability
        4. Flagged items (blanks, multis, orphans, corrections)
        5. Scoring parameters (for reproducibility)
    """
    from datetime import datetime
    import os

    ws = wb.create_sheet("Summary", 0)
    n_versions = len(versions)
    is_multi_version = n_versions > 1

    # ------------------------------------------------------------------ #
    # 1. Title
    # ------------------------------------------------------------------ #
    ws['A1'] = "MarkShark Exam Report"
    ws['A1'].font = Font(size=16, bold=True)

    # ------------------------------------------------------------------ #
    # 2. Project metadata
    # ------------------------------------------------------------------ #
    row = 3
    if project_name:
        ws[f'A{row}'] = "Project:"
        ws[f'A{row}'].font = FONT_BOLD
        ws[f'B{row}'] = project_name
        row += 1

    if input_csv_path:
        ws[f'A{row}'] = "Scores File:"
        ws[f'A{row}'].font = FONT_BOLD
        ws[f'B{row}'] = str(input_csv_path)
        row += 1

    ws[f'A{row}'] = "Generated:"
    ws[f'A{row}'].font = FONT_BOLD
    ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row += 1

    if corrections_applied > 0:
        ws[f'A{row}'] = "Corrections Applied:"
        ws[f'A{row}'].font = FONT_BOLD
        ws[f'B{row}'] = f"{corrections_applied} manual corrections from review"
        ws[f'B{row}'].font = Font(color="0000FF", italic=True)
        row += 1

    row += 1  # spacer

    # ------------------------------------------------------------------ #
    # 3. Per-version & amalgamated statistics + reliability
    # ------------------------------------------------------------------ #
    if is_multi_version:
        ws[f'A{row}'] = "Per-Version Statistics"
        ws[f'A{row}'].font = Font(size=13, bold=True)
        row += 1

        ver_headers = ["", "N Students", "Total Points", "Mean Score", "Mean %",
                        "Std Dev", "KR-20", "KR-21"]
        for col_idx, hdr in enumerate(ver_headers, start=1):
            ws.cell(row=row, column=col_idx, value=hdr)
        format_header_row(ws, row)
        row += 1

        for ver in versions:
            vs = version_stats[ver]
            ver_mean = vs['mean']
            total_pts = vs.get('total_points', k)
            ver_pct = (ver_mean / total_pts * 100) if total_pts > 0 else 0.0
            ver_kr20 = vs['kr20']
            ver_kr21 = vs['kr21']

            ws.cell(row=row, column=1, value=f"Version {ver}")
            ws.cell(row=row, column=1).font = FONT_BOLD
            ws.cell(row=row, column=2, value=vs['n_students'])
            ws.cell(row=row, column=3, value=int(total_pts) if total_pts == int(total_pts) else total_pts)
            ws.cell(row=row, column=4, value=f"{ver_mean:.2f}")
            ws.cell(row=row, column=5, value=f"{ver_pct:.1f}%")
            ws.cell(row=row, column=6, value=f"{vs['std']:.2f}")
            ws.cell(row=row, column=7, value=f"{ver_kr20:.3f}" if not np.isnan(ver_kr20) else "N/A")
            ws.cell(row=row, column=8, value=f"{ver_kr21:.3f}" if not np.isnan(ver_kr21) else "N/A")

            for col_idx in range(1, len(ver_headers) + 1):
                ws.cell(row=row, column=col_idx).border = BORDER_THIN
            row += 1

        row += 1  # spacer

    # ---- Amalgamated / overall stats ----
    if is_multi_version:
        ws[f'A{row}'] = "Amalgamated Exam Statistics (All Versions)"
    else:
        ws[f'A{row}'] = "Overall Exam Statistics"
    ws[f'A{row}'].font = Font(size=13, bold=True)
    row += 1

    max_total_pts = max(
        (vs.get('total_points', k) for vs in version_stats.values()), default=float(k)
    )
    max_pts_display = int(max_total_pts) if max_total_pts == int(max_total_pts) else max_total_pts

    stats = [
        ("Number of Students", n_students),
        ("Number of Questions", k),
        ("Total Points Possible", max_pts_display),
        ("Number of Versions", n_versions),
        ("Mean Score", f"{mean_total:.2f}"),
        ("Mean Percentage", f"{mean_total/max_total_pts*100:.1f}%" if max_total_pts > 0 else "N/A"),
        ("Standard Deviation", f"{std_total:.2f}"),
        ("KR-20 Reliability", f"{kr20_val:.3f}" if not np.isnan(kr20_val) else "N/A"),
        ("KR-21 Reliability", f"{kr21_val:.3f}" if not np.isnan(kr21_val) else "N/A"),
    ]
    for stat_name, stat_value in stats:
        ws[f'A{row}'] = stat_name
        ws[f'B{row}'] = stat_value
        row += 1

    # Reliability interpretation
    row += 1
    ws[f'A{row}'] = "Reliability Interpretation"
    ws[f'A{row}'].font = FONT_BOLD
    row += 1
    if not np.isnan(kr20_val):
        if kr20_val >= 0.80:
            ws[f'A{row}'] = "Excellent reliability (\u22650.80)"
            ws[f'A{row}'].fill = COLOR_GOOD
        elif kr20_val >= 0.70:
            ws[f'A{row}'] = "Good reliability (0.70-0.80)"
            ws[f'A{row}'].fill = COLOR_GOOD
        elif kr20_val >= 0.60:
            ws[f'A{row}'] = "Acceptable reliability (0.60-0.70)"
            ws[f'A{row}'].fill = COLOR_WARNING
        else:
            ws[f'A{row}'] = "Poor reliability (<0.60) - exam needs work"
            ws[f'A{row}'].fill = COLOR_PROBLEM

    # ------------------------------------------------------------------ #
    # 4. Absent Students & Flagged Items
    # ------------------------------------------------------------------ #
    row += 2
    ws[f'A{row}'] = "Absent Students & Flagged Items"
    ws[f'A{row}'].font = Font(size=13, bold=True)
    row += 1

    row = _write_flagged_items_section(
        ws, row, df, item_cols,
        orphan_scans, absent_students,
        corrections_detail, corrections_applied,
        auto_detected_students=auto_detected_students or [],
        versions=versions,
    )

    # ------------------------------------------------------------------ #
    # 5. Scoring Parameters (for reproducibility)
    # ------------------------------------------------------------------ #
    if scoring_params:
        row += 1
        ws[f'A{row}'] = "Scoring Parameters"
        ws[f'A{row}'].font = Font(size=13, bold=True)
        row += 1
        ws[f'A{row}'] = "(Recorded so results can be replicated if needed.)"
        ws[f'A{row}'].font = Font(italic=True, color="888888")
        row += 1

        param_headers = ["Parameter", "Value"]
        for col_idx, hdr in enumerate(param_headers, start=1):
            ws.cell(row=row, column=col_idx, value=hdr)
        format_header_row(ws, row)
        row += 1

        # Display-friendly names for common scoring params
        _labels = {
            'min_fill': 'Min Fill %',
            'fixed_thresh': 'Fixed Threshold',
            'auto_calibrate_thresh': 'Auto Calibrate Threshold',
            'calibrate_background': 'Background Calibration',
            'background_percentile': 'Background Percentile',
            'adaptive_rescoring': 'Adaptive Rescoring',
            'adaptive_max_adjustment': 'Adaptive Max Adjustment',
            'adaptive_min_above_floor': 'Adaptive Min Above Floor',
            'dpi': 'Render DPI',
        }
        for param_key, param_val in scoring_params.items():
            label = _labels.get(param_key, param_key)
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=str(param_val))
            for c in (1, 2):
                ws.cell(row=row, column=c).border = BORDER_THIN
            row += 1

    auto_size_columns(ws)


def _write_flagged_items_section(
    ws, start_row, df, item_cols,
    orphan_scans, absent_students,
    corrections_detail, corrections_applied,
    auto_detected_students=None,
    versions=None,
):
    """Write the inline flagged-items section on the Summary tab.

    Includes: blanks, multis, orphan scans, absent students, corrections.
    Returns the next available row number.
    """
    if df is None or item_cols is None:
        return start_row

    row = start_row

    # ---- Resolve column names ----
    id_col = _find_col(df, ['studentid', 'StudentID', 'Student_ID', 'student_id', 'ID', 'id'])
    last_col = _find_col(df, ['lastname', 'LastName', 'Last_Name', 'last_name', 'Last'])
    first_col = _find_col(df, ['firstname', 'FirstName', 'First_Name', 'first_name', 'First'])
    blank_col = _find_col(df, ['blank', 'Blank'])
    multi_col = _find_col(df, ['multi', 'Multi'])
    version_col = _find_col(df, ['version', 'Version'])
    flagdetails_col = _find_col(df, ['flagdetails', 'FlagDetails', 'flag_details'])

    # ---- Build corrections lookup ----
    corr_lookup: Dict[str, List[str]] = {}
    if corrections_detail is not None and not corrections_detail.empty:
        sid_cname = 'Student ID' if 'Student ID' in corrections_detail.columns else _find_col(
            corrections_detail, ['student_id', 'Student ID', 'StudentID']
        )
        q_cname = 'Question' if 'Question' in corrections_detail.columns else _find_col(
            corrections_detail, ['question', 'Question']
        )
        ans_cname = 'Corrected Answer' if 'Corrected Answer' in corrections_detail.columns else _find_col(
            corrections_detail, ['corrected_answer', 'Corrected Answer']
        )
        orig_cname = 'Original Answer' if 'Original Answer' in corrections_detail.columns else _find_col(
            corrections_detail, ['original_answer', 'Original Answer']
        )
        if sid_cname and q_cname and ans_cname:
            for _, crow in corrections_detail.iterrows():
                sid = _normalize_id(crow.get(sid_cname, ''))
                q = str(crow.get(q_cname, '')).strip()
                a = str(crow.get(ans_cname, '')).strip()
                orig = str(crow.get(orig_cname, '')).strip() if orig_cname else ''
                if orig.lower() in ('', 'nan', 'none'):
                    orig = ''
                if sid and q:
                    if q.upper() == 'ID':
                        # ID correction: show oldID → newID
                        # Key by the NEW id (corrected answer) since merge_corrections
                        # already updated the DataFrame's ID column.
                        new_id = _normalize_id(a)
                        label = f"ID: {sid}\u2192{a}"
                        corr_lookup.setdefault(new_id, []).append(label)
                    elif orig:
                        # Answer correction with original: Q15: A→C
                        corr_lookup.setdefault(sid, []).append(f"{q}: {orig}\u2192{a}")
                    else:
                        # Answer correction without original (legacy): Q15→C
                        corr_lookup.setdefault(sid, []).append(f"{q}\u2192{a}")

    # ---- Build orphan lookup (by scanned ID) for fast access ----
    orphan_lookup: Dict[str, dict] = {}
    if orphan_scans:
        for o in orphan_scans:
            oid = _normalize_id(o.get('ScannedID', ''))
            if oid:
                orphan_lookup[oid] = o

    # ---- Filter to student rows only (exclude KEY and VALUE rows) ----
    non_student_mask = df.apply(
        lambda r: any(str(cell).strip().upper() in ('KEY', 'VALUE') for cell in r), axis=1
    )
    students = df[~non_student_mask].copy()

    # ---- Collect flagged rows ----
    flagged_rows = []
    for idx, srow in students.iterrows():
        issues = []
        blank_n = multi_n = 0

        if blank_col:
            try:
                blank_n = int(srow.get(blank_col, 0))
            except (ValueError, TypeError):
                blank_n = 0
        if multi_col:
            try:
                multi_n = int(srow.get(multi_col, 0))
            except (ValueError, TypeError):
                multi_n = 0

        if blank_n > 0:
            issues.append(f"{blank_n} blank")
        if multi_n > 0:
            issues.append(f"{multi_n} multi")

        # Check for auto-detected version
        if version_col:
            raw_ver = str(srow.get(version_col, '')).strip()
            if raw_ver.endswith("*"):
                issues.append("version auto-detected")

        sid = _normalize_id(srow.get(id_col, '')) if id_col else ''

        # Check orphan status — from orphan_scans list OR FlagDetails column
        is_orphan = sid in orphan_lookup
        if not is_orphan and flagdetails_col:
            fd = str(srow.get(flagdetails_col, '')).strip()
            if 'orphan' in fd.lower():
                is_orphan = True
        orphan_info = orphan_lookup.get(sid) if is_orphan else None
        if is_orphan:
            match_type = orphan_info.get('MatchType', 'no_match') if orphan_info else 'no_match'
            possible = orphan_info.get('PossibleMatch', '') if orphan_info else ''
            if possible:
                issues.append(f"orphan (possible: {possible})")
            else:
                issues.append("orphan")

        # Corrections
        corr_list = corr_lookup.get(sid, [])
        if corr_list:
            issues.append(f"{len(corr_list)} correction(s)")

        if not issues:
            continue

        # Identify problem questions
        problem_qs = []
        for col in item_cols:
            answer = str(srow.get(col, '')).strip().upper() if pd.notna(srow.get(col)) else ''
            if not answer or answer in ('', 'NAN', 'BLANK', 'NONE', '?', '-'):
                problem_qs.append(f"{col}=BLANK")
            elif ',' in answer or answer == 'MULTI' or (len(answer) > 1 and answer not in ('BLANK', 'NONE', 'MULTI', 'NAN')):
                problem_qs.append(f"{col}={answer}")

        flagged_rows.append({
            'student_id': sid,
            'last_name': str(srow.get(last_col, '')).strip() if last_col else '',
            'first_name': str(srow.get(first_col, '')).strip() if first_col else '',
            'version': str(srow.get(version_col, '')).strip() if version_col else '',
            'issues': "; ".join(issues),
            'problem_questions': ", ".join(problem_qs[:10]) + ("..." if len(problem_qs) > 10 else ""),
            'corrections_applied': "; ".join(corr_list) if corr_list else "",
            'is_orphan': is_orphan,
        })

    # ---- Absent students (on roster but no scan) ----
    if absent_students:
        ws.cell(row=row, column=1, value=f"\u26a0 {len(absent_students)} absent student(s)")
        ws.cell(row=row, column=1).fill = COLOR_WARNING
        ws.cell(row=row, column=1).font = FONT_BOLD
        row += 1

        abs_headers = ["Student ID", "Last Name", "First Name"]
        for col_idx, hdr in enumerate(abs_headers, start=1):
            ws.cell(row=row, column=col_idx, value=hdr)
        format_header_row(ws, row)
        row += 1

        for student in absent_students:
            cell = ws.cell(row=row, column=1, value=student['StudentID'])
            cell.number_format = '@'
            ws.cell(row=row, column=2, value=student['LastName'])
            ws.cell(row=row, column=3, value=student['FirstName'])
            row += 1

        row += 1  # spacer after absent students

    # ---- Version Not Marked (auto-detected) ----
    if auto_detected_students and versions and len(versions) > 1:
        n_auto = len(auto_detected_students)
        ws.cell(
            row=row, column=1,
            value=f"\u26a0 {n_auto} student(s) did not mark a version "
                  f"(assigned by best score match)",
        )
        ws.cell(row=row, column=1).fill = COLOR_WARNING
        ws.cell(row=row, column=1).font = FONT_BOLD
        row += 1

        # Header: Student ID | Last Name | First Name | Assigned | Version A | Version B | ...
        auto_headers = ["Student ID", "Last Name", "First Name", "Assigned"]
        for ver in versions:
            auto_headers.append(f"Version {ver}")
        for col_idx, hdr in enumerate(auto_headers, start=1):
            ws.cell(row=row, column=col_idx, value=hdr)
        format_header_row(ws, row)
        row += 1

        total_q = auto_detected_students[0]['total'] if auto_detected_students else 0
        for student in auto_detected_students:
            cell = ws.cell(row=row, column=1, value=student['student_id'])
            cell.number_format = '@'
            ws.cell(row=row, column=2, value=student['last_name'])
            ws.cell(row=row, column=3, value=student['first_name'])
            ws.cell(row=row, column=4, value=student['auto_version'])
            ws.cell(row=row, column=4).font = FONT_BOLD

            for v_idx, ver in enumerate(versions):
                score = student['scores'].get(ver, 0)
                col_num = 5 + v_idx
                ws.cell(row=row, column=col_num, value=f"{score}/{total_q}")
                # Highlight the assigned version's score in green
                if ver == student['auto_version']:
                    ws.cell(row=row, column=col_num).fill = COLOR_GOOD
                    ws.cell(row=row, column=col_num).font = FONT_BOLD
            row += 1

        row += 1  # spacer

    # ---- Write flagged items table ----
    headers = ['Student ID', 'Last Name', 'First Name', 'Version', 'Issues',
               'Problem Questions', 'Corrections Applied']
    for col_idx, hdr in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=hdr)
    format_header_row(ws, row)
    row += 1

    if not flagged_rows:
        ws.cell(row=row, column=1, value="No flagged items.")
        ws.cell(row=row, column=1).font = Font(italic=True, color="666666")
        row += 1
    else:
        for item in flagged_rows:
            cell = ws.cell(row=row, column=1, value=item['student_id'])
            cell.number_format = '@'
            ws.cell(row=row, column=2, value=item['last_name'])
            ws.cell(row=row, column=3, value=item['first_name'])
            ws.cell(row=row, column=4, value=item['version'])
            ws.cell(row=row, column=5, value=item['issues'])
            ws.cell(row=row, column=6, value=item['problem_questions'])
            corr_cell = ws.cell(row=row, column=7, value=item['corrections_applied'])
            if item['corrections_applied']:
                corr_cell.font = Font(color="0000FF", italic=True)

            # Color-code by severity
            if item.get('is_orphan'):
                ws.cell(row=row, column=5).fill = COLOR_ORPHAN
            elif 'correction' in item['issues']:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).border = BORDER_THIN
            elif 'multi' in item['issues']:
                ws.cell(row=row, column=5).fill = COLOR_MULTI
            elif 'blank' in item['issues']:
                ws.cell(row=row, column=5).fill = COLOR_BLANK

            row += 1

        # Flagged summary line
        row += 1
        ws.cell(row=row, column=1, value=f"Total flagged: {len(flagged_rows)}")
        ws.cell(row=row, column=1).font = FONT_BOLD
        row += 1

    return row


def create_version_tab(
    wb, df_full, students_df, version, item_cols, key_series,
    difficulty, pb_vals, roster_df, orphan_scans
):
    """Create a tab for a specific exam version."""
    ws = wb.create_sheet(f"Version {version}")

    # Filter for this version (handle both 'Version' and 'version')
    version_col = 'version' if 'version' in df_full.columns else 'Version' if 'Version' in df_full.columns else None

    if version_col:
        # Match both "B" and "B*" (auto-detected) against base version "B"
        version_mask = df_full[version_col].astype(str).str.strip().str.rstrip("*") == str(version).strip()
        df_version = df_full[version_mask].copy()
    else:
        df_version = df_full.copy()

    # Get KEY row for this version
    key_row_data = df_version[df_version.apply(
        lambda row: any(str(cell).strip().upper() == 'KEY' for cell in row), axis=1
    )]

    # Get VALUE row for this version (point values per question)
    value_row_data = df_version[df_version.apply(
        lambda row: any(str(cell).strip().upper() == 'VALUE' for cell in row), axis=1
    )]

    # Get student rows (exclude KEY and VALUE rows)
    non_student_mask = df_version.apply(
        lambda row: any(str(cell).strip().upper() in ('KEY', 'VALUE') for cell in row), axis=1
    )
    student_rows = df_version[~non_student_mask]

    # Determine columns to display in the desired order:
    # LastName, FirstName, StudentID, Issue, correct, incorrect, blank, multi, percent, Version, Q1, Q2, ...
    display_cols = []

    # Identity columns first (check both cases)
    for col_variants in [('lastname', 'LastName'), ('firstname', 'FirstName'), ('studentid', 'StudentID')]:
        for variant in col_variants:
            if variant in df_version.columns:
                display_cols.append(variant)
                break

    # Add Issue column (will be computed)
    display_cols.append('Issue')

    # Score columns (check both cases)
    for col in ['score', 'correct', 'incorrect', 'blank', 'multi', 'percent']:
        if col in df_version.columns:
            display_cols.append(col)
        elif col.capitalize() in df_version.columns:
            display_cols.append(col.capitalize())

    # Version column (check both cases)
    if 'version' in df_version.columns:
        display_cols.append('version')
    elif 'Version' in df_version.columns:
        display_cols.append('Version')

    # Question columns
    display_cols.extend(item_cols)

    # Build column index map for quick lookup
    col_idx_map = {col: idx + 1 for idx, col in enumerate(display_cols)}

    # Get KEY answers for this version
    key_answers = {}
    if not key_row_data.empty:
        key_row = key_row_data.iloc[0]
        for col in item_cols:
            key_answers[col] = str(key_row.get(col, '')).strip().upper()

    # Get point values per question from VALUE row (default 1)
    point_values: Dict[str, float] = {}
    if not value_row_data.empty:
        val_row = value_row_data.iloc[0]
        for col in item_cols:
            try:
                point_values[col] = float(val_row.get(col, 1))
            except (ValueError, TypeError):
                point_values[col] = 1.0
    else:
        for col in item_cols:
            point_values[col] = 1.0

    total_pts = sum(point_values.values())
    total_pts_display = int(total_pts) if total_pts == int(total_pts) else total_pts

    # Write header (with enriched Score column name)
    for col_idx, col_name in enumerate(display_cols, start=1):
        header_text = col_name
        if col_name.lower() == 'score':
            header_text = f"Score (out of {total_pts_display})"
        ws.cell(row=1, column=col_idx, value=header_text)
    format_header_row(ws, 1)

    # Write student rows
    row_num = 2
    for _, student_row in student_rows.iterrows():
        # Determine issues for this student
        issues = []

        # Check for blank/multi answers
        blank_count = int(student_row.get('blank', 0))
        multi_count = int(student_row.get('multi', 0))
        if blank_count > 0:
            issues.append(f"{blank_count} blank")
        if multi_count > 0:
            issues.append(f"{multi_count} multi")

        # Check for auto-detected version (version bubble blank)
        if version_col:
            raw_ver = str(student_row.get(version_col, '')).strip()
            if raw_ver.endswith("*"):
                issues.append("version auto-detected")

        # Check roster matching (if available)
        if roster_df is not None:
            student_id = str(student_row.get('StudentID', '')).strip()
            if orphan_scans:
                # Check if this student is an orphan
                for orphan in orphan_scans:
                    if str(orphan.get('ScannedID', '')).strip() == student_id:
                        match_type = orphan.get('MatchType', 'no_match')
                        if match_type == 'no_match':
                            issues.append("ID mismatch")
                        elif match_type == 'probable':
                            issues.append("Fuzzy match")
                        break

        issue_text = "; ".join(issues) if issues else ""

        # Write all column values
        for col_idx, col_name in enumerate(display_cols, start=1):
            if col_name == 'Issue':
                cell = ws.cell(row=row_num, column=col_idx, value=issue_text)
                if issue_text:
                    cell.fill = COLOR_WARNING
            else:
                value = student_row.get(col_name, '')
                # Handle NaN/None from pandas (blank CSV cells become NaN)
                if pd.isna(value):
                    student_answer = ''
                else:
                    student_answer = str(value).strip().upper()

                # Check for blank or multi-answer cells in question columns
                if col_name in item_cols:
                    # Detect blank answers (empty, NaN, whitespace, or common blank indicators)
                    is_blank = (
                        not student_answer
                        or student_answer in ('', 'NAN', 'BLANK', 'NONE', '?')
                    )

                    # Detect multi-answer (contains comma, multiple letters, or "MULTI" indicator)
                    is_multi = (
                        not is_blank and (
                            ',' in student_answer or
                            student_answer == 'MULTI' or
                            (len(student_answer) > 1 and student_answer not in ('BLANK', 'NONE', 'MULTI', 'NAN'))
                        )
                    )

                    # Apply formatting based on answer type
                    if is_blank:
                        cell = ws.cell(row=row_num, column=col_idx, value="BLANK")
                        cell.fill = COLOR_BLANK
                        cell.alignment = Alignment(horizontal='center')
                    elif is_multi:
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal='center')
                        # Only colour multi-mark orange if it does NOT match
                        # a compound key (e.g. B&E).  Correct multis stay plain.
                        key_ans = key_answers.get(col_name, '')
                        if not key_ans or not _answer_matches_key(student_answer, key_ans):
                            cell.fill = COLOR_MULTI
                    else:
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal='center')
                        # Highlight incorrect answers in light red (only if not blank/multi)
                        if key_answers.get(col_name):
                            correct_answer = key_answers[col_name]
                            if student_answer and not _answer_matches_key(student_answer, correct_answer):
                                cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
                else:
                    # Non-question columns - just write the value
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    # Format StudentID column as text to prevent Excel treating it as a number
                    if col_name.lower() in ('studentid', 'student_id', 'id'):
                        cell.number_format = '@'

        row_num += 1

    # Add KEY answer row before statistics
    row_num += 1
    ws.cell(row=row_num, column=1, value="KEY")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    for col_name in item_cols:
        col_idx = col_idx_map.get(col_name)
        if col_idx and col_name in key_answers:
            cell = ws.cell(row=row_num, column=col_idx, value=key_answers[col_name])
            cell.alignment = Alignment(horizontal='center')
    format_key_row(ws, row_num)
    row_num += 1

    # Add VALUE row (max points per question)
    has_non_default = any(v != 1.0 for v in point_values.values())
    # Always write the VALUE row so teachers see it, even if all 1s
    ws.cell(row=row_num, column=1, value=f"Question Value (total points possible {total_pts_display})")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    for col_name in item_cols:
        col_idx = col_idx_map.get(col_name)
        if col_idx:
            pv = point_values.get(col_name, 1.0)
            cell = ws.cell(row=row_num, column=col_idx, value=int(pv) if pv == int(pv) else pv)
            cell.alignment = Alignment(horizontal='center')
    format_key_row(ws, row_num)
    row_num += 1

    # Add item statistics rows
    ws.cell(row=row_num, column=1, value="% Correct")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    for col_name in item_cols:
        col_idx = col_idx_map.get(col_name)
        if col_idx:
            pct = difficulty[col_name] * 100 if not np.isnan(difficulty[col_name]) else 0
            cell = ws.cell(row=row_num, column=col_idx, value=f"{pct:.1f}%")
            cell.alignment = Alignment(horizontal='center')
    row_num += 1

    ws.cell(row=row_num, column=1, value="Point-Biserial")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    for col_name in item_cols:
        col_idx = col_idx_map.get(col_name)
        if col_idx:
            pb = pb_vals[col_name]
            if not np.isnan(pb):
                cell = ws.cell(row=row_num, column=col_idx, value=f"{pb:.3f}")
                cell.alignment = Alignment(horizontal='center')
                # Color code based on quality
                if pb >= 0.20:
                    cell.fill = COLOR_GOOD
                elif pb >= 0.10:
                    cell.fill = COLOR_WARNING
                else:
                    cell.fill = COLOR_PROBLEM
    row_num += 1

    ws.cell(row=row_num, column=1, value="Item Quality")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    for col_name in item_cols:
        col_idx = col_idx_map.get(col_name)
        if col_idx:
            pb = pb_vals[col_name]
            if not np.isnan(pb):
                if pb >= 0.20:
                    quality = "✓ Good"
                    fill = COLOR_GOOD
                elif pb >= 0.10:
                    quality = "⚠ Review"
                    fill = COLOR_WARNING
                else:
                    quality = "✗ Problem"
                    fill = COLOR_PROBLEM
                cell = ws.cell(row=row_num, column=col_idx, value=quality)
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')
    row_num += 1

    # ---- Per-version Response Summary ----
    row_num += 2
    ws.cell(row=row_num, column=1, value="Response Distribution")
    ws.cell(row=row_num, column=1).font = Font(size=13, bold=True)
    row_num += 1

    n_ver_students = len(student_rows)

    # Collect all answer options seen for this version's students
    all_options = set()
    for col in item_cols:
        vals = student_rows[col].dropna().astype(str).str.strip().str.upper()
        for v in vals:
            if v and v not in ('', 'NAN', 'NONE', 'BLANK', '?', '-', 'MULTI'):
                if ',' not in v and len(v) <= 2:
                    all_options.add(v)
    sorted_options = sorted(all_options)

    # Build correct answers from this version's key
    correct_answers = {}
    if key_series is not None:
        for col in item_cols:
            if col in key_series.index:
                key_val = str(key_series[col]).strip().upper()
                if key_val and key_val not in ('', 'NAN', 'NONE'):
                    correct_answers[col] = key_val

    # Response summary header
    resp_headers = ['Question', 'Correct', '# Students', '% Correct']
    resp_headers += sorted_options
    resp_headers += ['Blank', 'Multi']
    for col_idx, hdr in enumerate(resp_headers, start=1):
        ws.cell(row=row_num, column=col_idx, value=hdr)
    format_header_row(ws, row_num)
    row_num += 1

    # Per-question response data
    for col_name in item_cols:
        answers = student_rows[col_name].fillna('').astype(str).str.strip().str.upper()

        counts = {}
        blank_count = 0
        multi_count = 0
        correct_count = 0
        correct_answer = correct_answers.get(col_name, '')

        # Determine if this key expects multiple bubbles (AND/partial)
        _is_compound_key = bool(correct_answer and any(op in correct_answer for op in ('&', '@', '~')))

        for answer in answers:
            if not answer or answer in ('', 'NAN', 'NONE', 'BLANK', '?', '-'):
                blank_count += 1
            elif _answer_matches_key(answer, correct_answer):
                correct_count += 1
                # For compound keys, bucket under the key representation
                if _is_compound_key:
                    counts[correct_answer] = counts.get(correct_answer, 0) + 1
                else:
                    counts[answer] = counts.get(answer, 0) + 1
            elif ',' in answer or answer == 'MULTI' or (len(answer) > 1 and answer not in sorted_options):
                multi_count += 1
            else:
                counts[answer] = counts.get(answer, 0) + 1

        pct_correct = (correct_count / n_ver_students * 100) if n_ver_students > 0 else 0

        ws.cell(row=row_num, column=1, value=col_name)
        ws.cell(row=row_num, column=1).font = FONT_BOLD
        ws.cell(row=row_num, column=2, value=correct_answer)
        ws.cell(row=row_num, column=3, value=n_ver_students)

        pct_cell = ws.cell(row=row_num, column=4, value=f"{pct_correct:.1f}%")
        if pct_correct >= 80:
            pct_cell.fill = COLOR_GOOD
        elif pct_correct >= 50:
            pct_cell.fill = COLOR_WARNING
        else:
            pct_cell.fill = COLOR_PROBLEM

        # Write counts for each option
        for opt_idx, option in enumerate(sorted_options):
            col_offset = 5 + opt_idx
            count = counts.get(option, 0)
            cell = ws.cell(row=row_num, column=col_offset, value=count)
            if option == correct_answer:
                cell.font = Font(bold=True, color="006600")
            elif count > 0 and n_ver_students > 0 and (count / n_ver_students) > 0.25:
                cell.fill = COLOR_WARNING

        # Blank and Multi columns
        blank_offset = 5 + len(sorted_options)
        multi_offset = blank_offset + 1
        blank_cell = ws.cell(row=row_num, column=blank_offset, value=blank_count)
        if blank_count > 0:
            blank_cell.fill = COLOR_BLANK
        multi_cell = ws.cell(row=row_num, column=multi_offset, value=multi_count)
        if multi_count > 0:
            multi_cell.fill = COLOR_MULTI

        row_num += 1

    # Legend
    row_num += 1
    ws.cell(row=row_num, column=1, value="Legend:")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    row_num += 1
    ws.cell(row=row_num, column=1, value="Green = ≥80% correct (easy)")
    ws.cell(row=row_num, column=1).fill = COLOR_GOOD
    row_num += 1
    ws.cell(row=row_num, column=1, value="Yellow = 50-80% correct (moderate)")
    ws.cell(row=row_num, column=1).fill = COLOR_WARNING
    row_num += 1
    ws.cell(row=row_num, column=1, value="Red = <50% correct (difficult)")
    ws.cell(row=row_num, column=1).fill = COLOR_PROBLEM
    row_num += 1
    ws.cell(row=row_num, column=1, value="Yellow count = popular wrong answer (>25% chose it)")

    auto_size_columns(ws)


def create_class_scores_tab(wb, df_full, item_cols, k):
    """
    Create a "Class Scores" tab with all students sorted alphabetically.

    This provides a simple roster view suitable for pasting into gradebooks:
    - LastName, FirstName, StudentID, Score, Correct, Percent, Version
    - Sorted by LastName first, then FirstName
    """
    ws = wb.create_sheet("Class Scores")

    # Determine column names (handle case variations) - do this first for filtering
    lastname_col = 'lastname' if 'lastname' in df_full.columns else 'LastName' if 'LastName' in df_full.columns else None
    firstname_col = 'firstname' if 'firstname' in df_full.columns else 'FirstName' if 'FirstName' in df_full.columns else None
    studentid_col = 'studentid' if 'studentid' in df_full.columns else 'StudentID' if 'StudentID' in df_full.columns else None
    score_col = 'score' if 'score' in df_full.columns else 'Score' if 'Score' in df_full.columns else None
    correct_col = 'correct' if 'correct' in df_full.columns else 'Correct' if 'Correct' in df_full.columns else None
    percent_col = 'percent' if 'percent' in df_full.columns else 'Percent' if 'Percent' in df_full.columns else None
    version_col = 'version' if 'version' in df_full.columns else 'Version' if 'Version' in df_full.columns else None

    # Filter out KEY/VALUE rows — same strategy used everywhere else in report_tools.
    # Checks ALL columns so any future marker rows are caught automatically.
    non_student_mask = df_full.apply(
        lambda row: any(str(cell).strip().upper() in ('KEY', 'VALUE') for cell in row), axis=1
    )
    df_students = df_full[~non_student_mask].copy()

    # Sort by last name first, then first name (case-insensitive)
    sort_cols = []
    if lastname_col:
        df_students['_sort_last'] = df_students[lastname_col].astype(str).str.upper()
        sort_cols.append('_sort_last')
    if firstname_col:
        df_students['_sort_first'] = df_students[firstname_col].astype(str).str.upper()
        sort_cols.append('_sort_first')

    if sort_cols:
        df_students = df_students.sort_values(sort_cols)
        # Drop sort columns
        df_students = df_students.drop(columns=[c for c in sort_cols if c in df_students.columns])

    # Write header
    headers = ['Last Name', 'First Name', 'Student ID', 'Score', 'Correct', 'Percent', 'Version']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    format_header_row(ws, 1)

    # Write student rows
    row_num = 2
    for _, row in df_students.iterrows():
        # Last Name
        ws.cell(row=row_num, column=1, value=row.get(lastname_col, '') if lastname_col else '')

        # First Name
        ws.cell(row=row_num, column=2, value=row.get(firstname_col, '') if firstname_col else '')

        # Student ID - format as text to prevent Excel treating it as a number
        cell = ws.cell(row=row_num, column=3, value=row.get(studentid_col, '') if studentid_col else '')
        cell.number_format = '@'

        # Score (points earned — may differ from correct count when questions have different point values)
        score_val = row.get(score_col, '') if score_col else ''
        ws.cell(row=row_num, column=4, value=score_val)

        # Correct (number of correctly answered questions)
        correct_val = row.get(correct_col, '') if correct_col else ''
        ws.cell(row=row_num, column=5, value=correct_val)

        # Percent
        percent = row.get(percent_col, '') if percent_col else ''
        if percent and str(percent).strip():
            try:
                pct_val = float(percent)
                ws.cell(row=row_num, column=6, value=f"{pct_val:.1f}%")
            except (ValueError, TypeError):
                ws.cell(row=row_num, column=6, value=percent)
        else:
            ws.cell(row=row_num, column=6, value='')

        # Version
        ws.cell(row=row_num, column=7, value=row.get(version_col, '') if version_col else '')

        row_num += 1

    # Add summary at bottom
    row_num += 1
    ws.cell(row=row_num, column=1, value="Total Students:")
    ws.cell(row=row_num, column=1).font = FONT_BOLD
    ws.cell(row=row_num, column=2, value=len(df_students))

    auto_size_columns(ws)


def create_answer_key_tab(wb, item_cols, versions, version_stats):
    """
    Create an "Answer Key" tab showing the correct answer for each question
    across all versions.

    Layout:
        Question | Version A | Version B | ...
        Q1       | C         | A         | ...
        Q2       | B         | D         | ...
        ...
    """
    ws = wb.create_sheet("Answer Key")

    # Header row
    headers = ["Question"] + [f"Version {v}" for v in versions]
    for col_idx, hdr in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=hdr)
    format_header_row(ws, 1)

    # One row per question
    for q_idx, q_col in enumerate(item_cols):
        row_num = q_idx + 2
        ws.cell(row=row_num, column=1, value=q_col)
        ws.cell(row=row_num, column=1).font = FONT_BOLD

        for v_idx, ver in enumerate(versions):
            ks = version_stats[ver].get('key_series')
            if ks is not None and q_col in ks.index:
                answer = str(ks[q_col]).strip().upper()
                if answer and answer not in ('NAN', 'NONE', ''):
                    ws.cell(row=row_num, column=v_idx + 2, value=answer)

        # Apply thin borders to all cells in the row
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=c).border = BORDER_THIN

    # Summary
    row_num = len(item_cols) + 3
    ws.cell(row=row_num, column=1, value=f"Total Questions: {len(item_cols)}")
    ws.cell(row=row_num, column=1).font = FONT_BOLD

    auto_size_columns(ws)


def create_flagged_items_tab(
    wb, df: pd.DataFrame, item_cols: List[str],
    corrections_detail: Optional[pd.DataFrame] = None,
):
    """
    Create a Flagged Items tab listing all students with issues.

    A student is flagged if they have blank answers, multi-answers, or
    a correction has been applied. This gives teachers a single view
    of everything that needs attention.

    The tab is inserted at position 0 (first tab in the workbook).
    """
    ws = wb.create_sheet("Flagged Items", 0)

    # Resolve column names
    id_col = _find_col(df, ['studentid', 'StudentID', 'Student_ID', 'student_id', 'ID', 'id'])
    last_col = _find_col(df, ['lastname', 'LastName', 'Last_Name', 'last_name', 'Last'])
    first_col = _find_col(df, ['firstname', 'FirstName', 'First_Name', 'first_name', 'First'])
    blank_col = _find_col(df, ['blank', 'Blank'])
    multi_col = _find_col(df, ['multi', 'Multi'])
    version_col = _find_col(df, ['version', 'Version'])

    # Build corrections lookup: student_id -> list of {question, corrected_answer}
    corr_lookup: Dict[str, List[str]] = {}
    if corrections_detail is not None and not corrections_detail.empty:
        sid_cname = 'Student ID' if 'Student ID' in corrections_detail.columns else _find_col(
            corrections_detail, ['student_id', 'Student ID', 'StudentID']
        )
        q_cname = 'Question' if 'Question' in corrections_detail.columns else _find_col(
            corrections_detail, ['question', 'Question']
        )
        ans_cname = 'Corrected Answer' if 'Corrected Answer' in corrections_detail.columns else _find_col(
            corrections_detail, ['corrected_answer', 'Corrected Answer']
        )
        if sid_cname and q_cname and ans_cname:
            for _, crow in corrections_detail.iterrows():
                sid = _normalize_id(crow.get(sid_cname, ''))
                q = str(crow.get(q_cname, '')).strip()
                a = str(crow.get(ans_cname, '')).strip()
                if sid and q:
                    corr_lookup.setdefault(sid, []).append(f"{q}→{a}")

    # Filter to student rows only (exclude KEY and VALUE rows)
    non_student_mask = df.apply(
        lambda row: any(str(cell).strip().upper() in ('KEY', 'VALUE') for cell in row), axis=1
    )
    students = df[~non_student_mask].copy()

    # Determine which students are flagged
    flagged_rows = []
    for idx, row in students.iterrows():
        issues = []
        blank_n = 0
        multi_n = 0

        if blank_col:
            try:
                blank_n = int(row.get(blank_col, 0))
            except (ValueError, TypeError):
                blank_n = 0
        if multi_col:
            try:
                multi_n = int(row.get(multi_col, 0))
            except (ValueError, TypeError):
                multi_n = 0

        if blank_n > 0:
            issues.append(f"{blank_n} blank")
        if multi_n > 0:
            issues.append(f"{multi_n} multi")

        sid = _normalize_id(row.get(id_col, '')) if id_col else ''
        corr_list = corr_lookup.get(sid, [])
        if corr_list:
            issues.append(f"{len(corr_list)} correction(s)")

        if not issues:
            continue  # Not flagged

        # Identify problematic questions
        problem_qs = []
        for col in item_cols:
            answer = str(row.get(col, '')).strip().upper() if pd.notna(row.get(col)) else ''
            if not answer or answer in ('', 'NAN', 'BLANK', 'NONE', '?', '-'):
                problem_qs.append(f"{col}=BLANK")
            elif ',' in answer or answer == 'MULTI' or (len(answer) > 1 and answer not in ('BLANK', 'NONE', 'MULTI', 'NAN')):
                problem_qs.append(f"{col}={answer}")

        flagged_rows.append({
            'student_id': sid,
            'last_name': str(row.get(last_col, '')).strip() if last_col else '',
            'first_name': str(row.get(first_col, '')).strip() if first_col else '',
            'version': str(row.get(version_col, '')).strip() if version_col else '',
            'issues': "; ".join(issues),
            'problem_questions': ", ".join(problem_qs[:10]) + ("..." if len(problem_qs) > 10 else ""),
            'corrections_applied': "; ".join(corr_list) if corr_list else "",
        })

    # Write header
    headers = ['Student ID', 'Last Name', 'First Name', 'Version', 'Issues',
               'Problem Questions', 'Corrections Applied']
    for col_idx, hdr in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=hdr)
    format_header_row(ws, 1)

    if not flagged_rows:
        ws.cell(row=2, column=1, value="No flagged items found.")
        ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
        auto_size_columns(ws)
        return

    # Write data
    row_num = 2
    for item in flagged_rows:
        cell = ws.cell(row=row_num, column=1, value=item['student_id'])
        cell.number_format = '@'
        ws.cell(row=row_num, column=2, value=item['last_name'])
        ws.cell(row=row_num, column=3, value=item['first_name'])
        ws.cell(row=row_num, column=4, value=item['version'])
        ws.cell(row=row_num, column=5, value=item['issues'])
        ws.cell(row=row_num, column=6, value=item['problem_questions'])
        corr_cell = ws.cell(row=row_num, column=7, value=item['corrections_applied'])
        if item['corrections_applied']:
            corr_cell.font = Font(color="0000FF", italic=True)

        # Highlight row based on severity
        if 'correction' in item['issues']:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).border = BORDER_THIN
        elif 'multi' in item['issues']:
            ws.cell(row=row_num, column=5).fill = COLOR_MULTI
        elif 'blank' in item['issues']:
            ws.cell(row=row_num, column=5).fill = COLOR_BLANK

        row_num += 1

    # Summary
    row_num += 1
    ws.cell(row=row_num, column=1, value=f"Total flagged: {len(flagged_rows)}")
    ws.cell(row=row_num, column=1).font = FONT_BOLD

    auto_size_columns(ws)


