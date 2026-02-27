"""
LMS Integration page — import gradebook exports and map columns.

Two tabs:
1. Import & Map — load an LMS gradebook (CSV/TSV/XLSX), assign columns to
   MarkShark properties, save the mapping as a reusable filter, and export
   a MarkShark-compatible roster.
2. Write Scores — apply MarkShark results back into an LMS gradebook file.
"""

import csv
import io
from pathlib import Path

import openpyxl

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTabWidget,
    QComboBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QGroupBox,
    QFormLayout,
    QLineEdit,
    QMessageBox,
    QSpinBox,
    QInputDialog,
)

from ..widgets import FileSelector, PageHeader
from ..models.lms_filter_registry import LmsFilterRegistry
from ..utils import RUN_BUTTON_STYLE as _RUN_BTN_STYLE

_NONE_LABEL = "(none)"
_ADD_NEW_COL = "\u2795 Add a new column\u2026"

# Maximum preview rows shown in the table
_MAX_PREVIEW_ROWS = 8


class LmsIntegrationPage(QWidget):
    """LMS Integration page with tabbed sub-tools."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._registry = LmsFilterRegistry()
        self._headers: list[str] = []
        self._rows: list[list[str]] = []
        self._file_ext: str = ""
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        header = PageHeader(
            "MarkShark LMS Integration",
            "Import gradebook exports from your LMS and map columns to "
            "MarkShark properties. Save mappings as reusable filters.",
        )
        layout.addWidget(header)

        tabs = QTabWidget()
        tabs.addTab(self._build_import_tab(), "Import && Map Columns")
        tabs.addTab(self._build_scores_tab(), "Write Scores Back")
        layout.addWidget(tabs, 1)

    # ==================================================================
    # Tab 1: Import & Map Columns
    # ==================================================================

    def _build_import_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)

        # ── File selection ──
        self.lms_file = FileSelector(
            "LMS gradebook:",
            "Spreadsheets (*.csv *.tsv *.tab *.xlsx *.xls)",
            "Select a gradebook export (CSV, TSV, XLSX)...",
        )
        self.lms_file.file_selected.connect(self._on_file_selected)
        layout.addWidget(self.lms_file)

        # ── Delimiter / skip rows (for CSV/TSV) ──
        opts_row = QHBoxLayout()
        opts_row.addWidget(QLabel("Delimiter:"))
        self.delim_combo = QComboBox()
        self.delim_combo.addItems(["Comma (,)", "Tab (\\t)", "Semicolon (;)", "Pipe (|)"])
        self.delim_combo.setToolTip("Delimiter used in the text file. Ignored for XLSX.")
        self.delim_combo.currentIndexChanged.connect(self._reload_preview)
        opts_row.addWidget(self.delim_combo)

        opts_row.addSpacing(20)
        opts_row.addWidget(QLabel("Skip rows:"))
        self.skip_rows_spin = QSpinBox()
        self.skip_rows_spin.setRange(0, 50)
        self.skip_rows_spin.setToolTip(
            "Number of non-header rows to skip at the top of the file "
            "(some LMS exports have extra info rows before the header)."
        )
        self.skip_rows_spin.valueChanged.connect(self._reload_preview)
        opts_row.addWidget(self.skip_rows_spin)
        opts_row.addStretch()

        reload_btn = QPushButton("Reload Preview")
        reload_btn.clicked.connect(self._reload_preview)
        opts_row.addWidget(reload_btn)
        layout.addLayout(opts_row)

        # ── Preview table ──
        self.preview_table = QTableWidget()
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.verticalHeader().setVisible(True)
        self.preview_table.setMinimumHeight(140)
        layout.addWidget(self.preview_table, 1)

        # ── Column mapping ──
        map_group = QGroupBox("Column Mapping")
        map_layout = QFormLayout(map_group)

        self.sid_combo = QComboBox()
        self.sid_combo.setMinimumWidth(220)
        map_layout.addRow("Student ID column:", self.sid_combo)

        self.last_combo = QComboBox()
        self.last_combo.setMinimumWidth(220)
        map_layout.addRow("Last Name column:", self.last_combo)

        self.first_combo = QComboBox()
        self.first_combo.setMinimumWidth(220)
        map_layout.addRow("First Name column:", self.first_combo)

        self.combined_combo = QComboBox()
        self.combined_combo.setMinimumWidth(220)
        self.combined_combo.setToolTip(
            'For LMS exports with a single "Last, First" name column.\n'
            "MarkShark will split on comma to get last and first names.\n"
            "Leave as (none) if your LMS has separate name columns."
        )
        map_layout.addRow("Combined Name column:", self.combined_combo)

        layout.addWidget(map_group)

        # ── Saved filters ──
        filter_group = QGroupBox("Saved Filters")
        filter_layout = QHBoxLayout(filter_group)

        self.filter_combo = QComboBox()
        self.filter_combo.setMinimumWidth(200)
        self._refresh_filter_list()
        filter_layout.addWidget(self.filter_combo, 1)

        load_btn = QPushButton("Load Filter")
        load_btn.clicked.connect(self._load_filter)
        filter_layout.addWidget(load_btn)

        save_btn = QPushButton("Save As...")
        save_btn.clicked.connect(self._save_filter)
        filter_layout.addWidget(save_btn)

        del_btn = QPushButton("Delete")
        del_btn.clicked.connect(self._delete_filter)
        filter_layout.addWidget(del_btn)

        layout.addWidget(filter_group)

        # ── Export roster button ──
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.import_status = QLabel("")
        btn_row.addWidget(self.import_status)

        self.export_roster_btn = QPushButton("Export as MarkShark Roster")
        self.export_roster_btn.setStyleSheet(_RUN_BTN_STYLE)
        self.export_roster_btn.clicked.connect(self._export_roster)
        btn_row.addWidget(self.export_roster_btn)
        layout.addLayout(btn_row)

        return tab

    # ------------------------------------------------------------------
    # File loading
    # ------------------------------------------------------------------

    def _get_delimiter(self) -> str:
        idx = self.delim_combo.currentIndex()
        return [",", "\t", ";", "|"][idx]

    def _on_file_selected(self):
        """Called when the user picks a new LMS file."""
        self._reload_preview()

    def _reload_preview(self):
        """(Re)parse the selected file and populate the preview table."""
        path = self.lms_file.path()
        if not path or not Path(path).is_file():
            return

        self._file_ext = Path(path).suffix.lower()
        skip = self.skip_rows_spin.value()

        try:
            if self._file_ext in (".xlsx", ".xls"):
                self._headers, self._rows = self._read_excel(path, skip)
            else:
                delim = self._get_delimiter()
                self._headers, self._rows = self._read_text(path, delim, skip)
        except Exception as e:
            QMessageBox.warning(self, "Parse Error", f"Failed to read file:\n{e}")
            self._headers, self._rows = [], []
            return

        self._populate_preview()
        self._populate_mapping_combos()

    @staticmethod
    def _read_text(path: str, delimiter: str, skip: int) -> tuple:
        """Read a CSV/TSV file. Returns (headers, rows)."""
        with open(path, newline="", encoding="utf-8-sig") as f:
            # Skip leading rows
            for _ in range(skip):
                next(f, None)
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, [])
            headers = [h.strip() for h in headers]
            rows = []
            for row in reader:
                rows.append([c.strip() for c in row])
                if len(rows) >= _MAX_PREVIEW_ROWS:
                    break
        return headers, rows

    @staticmethod
    def _read_excel(path: str, skip: int) -> tuple:
        """Read the first sheet of an XLSX/XLS file. Returns (headers, rows)."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(c) if c is not None else "" for c in row])
        wb.close()

        if skip >= len(all_rows):
            return [], []
        # First row after skip is headers
        headers = all_rows[skip]
        data = all_rows[skip + 1: skip + 1 + _MAX_PREVIEW_ROWS]
        return headers, data

    def _populate_preview(self):
        """Fill the preview QTableWidget with parsed data."""
        tbl = self.preview_table
        tbl.clear()
        if not self._headers:
            tbl.setRowCount(0)
            tbl.setColumnCount(0)
            return

        tbl.setColumnCount(len(self._headers))
        tbl.setHorizontalHeaderLabels(self._headers)
        tbl.setRowCount(len(self._rows))

        for r, row in enumerate(self._rows):
            for c, val in enumerate(row):
                if c < len(self._headers):
                    tbl.setItem(r, c, QTableWidgetItem(val))

        tbl.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )

    def _populate_mapping_combos(self):
        """Fill the mapping combo boxes with column names from the file."""
        choices = [_NONE_LABEL] + self._headers

        for combo in (self.sid_combo, self.last_combo, self.first_combo,
                      self.combined_combo):
            combo.blockSignals(True)
            combo.clear()
            combo.addItems(choices)
            combo.blockSignals(False)

        # Auto-detect common column names
        self._auto_detect_columns()

    def _auto_detect_columns(self):
        """Try to match common LMS column names to MarkShark properties."""
        lower_headers = {h.lower().strip(): h for h in self._headers}

        # Student ID aliases
        sid_aliases = [
            "student id", "studentid", "student_id", "sid", "sis user id",
            "sis id", "id", "user id", "userid", "student number",
            "student_number", "enrollment id",
        ]
        for alias in sid_aliases:
            if alias in lower_headers:
                idx = self._headers.index(lower_headers[alias])
                self.sid_combo.setCurrentIndex(idx + 1)  # +1 for (none)
                break

        # Last name aliases
        last_aliases = [
            "last name", "lastname", "last_name", "last", "surname",
            "family name", "family_name",
        ]
        for alias in last_aliases:
            if alias in lower_headers:
                idx = self._headers.index(lower_headers[alias])
                self.last_combo.setCurrentIndex(idx + 1)
                break

        # First name aliases
        first_aliases = [
            "first name", "firstname", "first_name", "first", "given name",
            "given_name", "preferred name",
        ]
        for alias in first_aliases:
            if alias in lower_headers:
                idx = self._headers.index(lower_headers[alias])
                self.first_combo.setCurrentIndex(idx + 1)
                break

        # Combined "Last, First" name column aliases
        combined_aliases = [
            "sortable name", "sortable_name", "student",
            "student name", "student_name", "name",
        ]
        for alias in combined_aliases:
            if alias in lower_headers:
                idx = self._headers.index(lower_headers[alias])
                self.combined_combo.setCurrentIndex(idx + 1)
                break

    # ------------------------------------------------------------------
    # Filter management
    # ------------------------------------------------------------------

    def _refresh_filter_list(self):
        """Reload the filter combo from the registry."""
        self.filter_combo.clear()
        names = self._registry.list_names()
        if names:
            self.filter_combo.addItems(names)
        else:
            self.filter_combo.addItem("(no saved filters)")

    def _save_filter(self):
        """Save the current column mapping as a named filter."""
        sid_col = self.sid_combo.currentText()
        last_col = self.last_combo.currentText()
        first_col = self.first_combo.currentText()

        if sid_col == _NONE_LABEL and last_col == _NONE_LABEL:
            QMessageBox.warning(
                self, "No Mapping",
                "Please assign at least one column before saving a filter.",
            )
            return

        # Default suggestion: filename without extension
        default_name = ""
        path = self.lms_file.path()
        if path:
            default_name = Path(path).stem

        name, ok = QInputDialog.getText(
            self, "Save Filter",
            "Filter name:",
            QLineEdit.EchoMode.Normal,
            default_name,
        )
        if not ok or not name.strip():
            return

        combined_col = self.combined_combo.currentText()
        self._registry.save_filter(
            name=name.strip(),
            student_id_col=sid_col if sid_col != _NONE_LABEL else "",
            last_name_col=last_col if last_col != _NONE_LABEL else "",
            first_name_col=first_col if first_col != _NONE_LABEL else "",
            delimiter=self._get_delimiter(),
            skip_rows=self.skip_rows_spin.value(),
            combined_name_col=combined_col if combined_col != _NONE_LABEL else "",
        )
        self._refresh_filter_list()
        self._refresh_scores_filter_list()
        # Select the newly saved filter
        idx = self.filter_combo.findText(name.strip())
        if idx >= 0:
            self.filter_combo.setCurrentIndex(idx)
        self.import_status.setText(f"Filter \"{name.strip()}\" saved.")

    def _load_filter(self):
        """Load a saved filter and apply its column mappings."""
        name = self.filter_combo.currentText()
        if not name or name == "(no saved filters)":
            return

        filt = self._registry.get_filter(name)
        if not filt:
            return

        # Apply delimiter
        delim_map = {",": 0, "\t": 1, ";": 2, "|": 3}
        delim_idx = delim_map.get(filt.get("delimiter", ","), 0)
        self.delim_combo.setCurrentIndex(delim_idx)

        # Apply skip rows
        self.skip_rows_spin.setValue(filt.get("skip_rows", 0))

        # If a file is loaded, reload with new settings then set combos
        if self.lms_file.path() and Path(self.lms_file.path()).is_file():
            self._reload_preview()

        # Set column combos (after reload so headers are populated)
        self._set_combo_by_text(self.sid_combo, filt.get("student_id_col", ""))
        self._set_combo_by_text(self.last_combo, filt.get("last_name_col", ""))
        self._set_combo_by_text(self.first_combo, filt.get("first_name_col", ""))
        self._set_combo_by_text(self.combined_combo, filt.get("combined_name_col", ""))

        self._registry.touch_last_used(name)
        self.import_status.setText(f"Filter \"{name}\" loaded.")

    def _delete_filter(self):
        """Delete the currently selected filter."""
        name = self.filter_combo.currentText()
        if not name or name == "(no saved filters)":
            return

        reply = QMessageBox.question(
            self, "Delete Filter",
            f"Delete filter \"{name}\"?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._registry.delete_filter(name)
            self._refresh_filter_list()
            self._refresh_scores_filter_list()
            self.import_status.setText(f"Filter \"{name}\" deleted.")

    @staticmethod
    def _set_combo_by_text(combo: QComboBox, text: str):
        """Set a combo box to the item matching text, or (none)."""
        if not text:
            combo.setCurrentIndex(0)
            return
        idx = combo.findText(text)
        if idx >= 0:
            combo.setCurrentIndex(idx)
        else:
            combo.setCurrentIndex(0)

    # ------------------------------------------------------------------
    # Export roster
    # ------------------------------------------------------------------

    def _export_roster(self):
        """Export the mapped columns as a MarkShark-compatible roster CSV."""
        path = self.lms_file.path()
        if not path or not Path(path).is_file():
            QMessageBox.warning(self, "No File", "Please select an LMS gradebook file first.")
            return

        sid_col = self.sid_combo.currentText()
        last_col = self.last_combo.currentText()
        first_col = self.first_combo.currentText()
        combined_col = self.combined_combo.currentText()

        if sid_col == _NONE_LABEL:
            QMessageBox.warning(
                self, "No Student ID",
                "Please assign the Student ID column before exporting.",
            )
            return

        # Read the full file (not just preview rows)
        try:
            headers, rows = self._read_full_file(path)
        except Exception as e:
            QMessageBox.critical(self, "Read Error", f"Failed to read file:\n{e}")
            return

        if not headers or not rows:
            QMessageBox.warning(self, "Empty File", "The file appears to be empty.")
            return

        # Find column indices
        sid_idx = headers.index(sid_col) if sid_col in headers else -1
        last_idx = headers.index(last_col) if last_col != _NONE_LABEL and last_col in headers else -1
        first_idx = headers.index(first_col) if first_col != _NONE_LABEL and first_col in headers else -1
        combined_idx = headers.index(combined_col) if combined_col != _NONE_LABEL and combined_col in headers else -1

        if sid_idx < 0:
            QMessageBox.warning(self, "Column Not Found", f"Column \"{sid_col}\" not found in file.")
            return

        # Build roster rows
        roster_rows = []
        for row in rows:
            sid = row[sid_idx] if sid_idx < len(row) else ""

            if combined_idx >= 0 and combined_idx < len(row):
                # Split "Last, First" combined name column
                raw = row[combined_idx].strip()
                if "," in raw:
                    last, first = raw.split(",", 1)
                    last = last.strip()
                    first = first.strip()
                else:
                    # No comma — treat the whole value as last name
                    last = raw
                    first = ""
            else:
                last = row[last_idx] if last_idx >= 0 and last_idx < len(row) else ""
                first = row[first_idx] if first_idx >= 0 and first_idx < len(row) else ""

            if sid.strip():
                roster_rows.append((sid.strip(), last.strip(), first.strip()))

        if not roster_rows:
            QMessageBox.warning(self, "No Students", "No student rows found with the selected mapping.")
            return

        # Ask for save location
        from PySide6.QtWidgets import QFileDialog
        save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Roster", "roster.csv", "CSV files (*.csv)"
        )
        if not save_path:
            return

        try:
            with open(save_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["StudentID", "LastName", "FirstName"])
                for sid, last, first in roster_rows:
                    writer.writerow([sid, last, first])
            self.import_status.setText(f"Roster exported — {len(roster_rows)} students.")
        except Exception as e:
            QMessageBox.critical(self, "Write Error", f"Failed to write roster:\n{e}")

    def _read_full_file(self, path: str) -> tuple:
        """Read all rows from the file (not limited to preview count)."""
        ext = Path(path).suffix.lower()
        skip = self.skip_rows_spin.value()

        if ext in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append([str(c) if c is not None else "" for c in row])
            wb.close()
            if skip >= len(all_rows):
                return [], []
            headers = all_rows[skip]
            data = all_rows[skip + 1:]
            return headers, data
        else:
            delim = self._get_delimiter()
            with open(path, newline="", encoding="utf-8-sig") as f:
                for _ in range(skip):
                    next(f, None)
                reader = csv.reader(f, delimiter=delim)
                headers = next(reader, [])
                headers = [h.strip() for h in headers]
                rows = [[c.strip() for c in row] for row in reader]
            return headers, rows

    # ==================================================================
    # Tab 2: Write Scores Back
    # ==================================================================

    def _build_scores_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)

        desc = QLabel(
            "Write MarkShark scores back into an LMS gradebook file. "
            "Select the original LMS export and the MarkShark exam report "
            "(which includes any corrections from the Review panel), "
            "then choose which score column to insert."
        )
        desc.setWordWrap(True)
        layout.addWidget(desc)

        self.scores_lms_file = FileSelector(
            "LMS gradebook:",
            "Spreadsheets (*.csv *.tsv *.tab *.xlsx *.xls)",
            "Select the original LMS gradebook...",
        )
        self.scores_lms_file.file_selected.connect(self._on_scores_lms_selected)
        layout.addWidget(self.scores_lms_file)

        self.scores_results = FileSelector(
            "Exam report:",
            "Excel files (*.xlsx)",
            "Select exam_report.xlsx from a scored project...",
        )
        layout.addWidget(self.scores_results)

        # ── Mapping options ──
        map_group = QGroupBox("Mapping Options")
        map_layout = QFormLayout(map_group)

        self.scores_filter_combo = QComboBox()
        self.scores_filter_combo.setMinimumWidth(200)
        self._refresh_scores_filter_list()
        map_layout.addRow("Use saved filter:", self.scores_filter_combo)

        self.scores_sid_combo = QComboBox()
        self.scores_sid_combo.setMinimumWidth(220)
        self.scores_sid_combo.setToolTip(
            "Which column in the LMS file contains the Student ID "
            "that matches MarkShark results."
        )
        map_layout.addRow("LMS Student ID column:", self.scores_sid_combo)

        self.scores_value_combo = QComboBox()
        self.scores_value_combo.setMinimumWidth(220)
        self.scores_value_combo.addItems(["Score (raw)", "Percent"])
        self.scores_value_combo.setToolTip(
            "Which value from MarkShark results to write into the LMS file."
        )
        map_layout.addRow("Value to write:", self.scores_value_combo)

        self.scores_target_combo = QComboBox()
        self.scores_target_combo.setMinimumWidth(220)
        self.scores_target_combo.setToolTip(
            "Choose an existing column to overwrite, or add a new one."
        )
        self.scores_target_combo.addItem(_ADD_NEW_COL)
        self.scores_target_combo.currentIndexChanged.connect(
            self._on_target_col_changed
        )
        map_layout.addRow("Target column:", self.scores_target_combo)

        self.absent_combo = QComboBox()
        self.absent_combo.setMinimumWidth(220)
        self.absent_combo.addItems(["Leave blank", "Enter zero (0)"])
        self.absent_combo.setToolTip(
            "For students with no matching score AND a blank target cell:\n"
            "'Leave blank' does nothing; 'Enter zero' writes 0.\n"
            "Cells that already have content are never overwritten."
        )
        map_layout.addRow("Absent students:", self.absent_combo)

        layout.addWidget(map_group)

        # ── Load filter button for scores tab ──
        filter_row = QHBoxLayout()
        filter_row.addStretch()
        load_filter_btn = QPushButton("Apply Filter to SID Column")
        load_filter_btn.clicked.connect(self._apply_filter_to_scores)
        filter_row.addWidget(load_filter_btn)
        layout.addLayout(filter_row)

        layout.addStretch()

        # ── Output ──
        self.scores_output = FileSelector(
            "Save output as:",
            "Spreadsheets (*.csv *.xlsx)",
            "Save updated gradebook as...",
            save_mode=True,
        )
        layout.addWidget(self.scores_output)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.scores_status = QLabel("")
        btn_row.addWidget(self.scores_status)
        self.scores_run_btn = QPushButton("Write Scores")
        self.scores_run_btn.setStyleSheet(_RUN_BTN_STYLE)
        self.scores_run_btn.clicked.connect(self._run_write_scores)
        btn_row.addWidget(self.scores_run_btn)
        layout.addLayout(btn_row)

        return tab

    def _refresh_scores_filter_list(self):
        """Reload the filter combo on the scores tab."""
        self.scores_filter_combo.clear()
        self.scores_filter_combo.addItem("(none — set manually)")
        for name in self._registry.list_names():
            self.scores_filter_combo.addItem(name)

    def _apply_filter_to_scores(self):
        """Apply a saved filter's SID column to the scores tab."""
        name = self.scores_filter_combo.currentText()
        if not name or name == "(none — set manually)":
            QMessageBox.information(
                self, "No Filter",
                "Select a saved filter first, or set the Student ID column manually.",
            )
            return

        filt = self._registry.get_filter(name)
        if not filt:
            return

        # We need to load the LMS file to populate the SID combo
        lms_path = self.scores_lms_file.path()
        if not lms_path or not Path(lms_path).is_file():
            QMessageBox.warning(self, "No File", "Please select the LMS gradebook file first.")
            return

        ext = Path(lms_path).suffix.lower()
        skip = filt.get("skip_rows", 0)
        try:
            if ext in (".xlsx", ".xls"):
                headers, _ = self._read_excel(lms_path, skip)
            else:
                delim = filt.get("delimiter", ",")
                headers, _ = self._read_text(lms_path, delim, skip)
        except Exception as e:
            QMessageBox.warning(self, "Parse Error", f"Failed to read file:\n{e}")
            return

        # Populate the SID combo and target column combo from headers
        self.scores_sid_combo.clear()
        self.scores_sid_combo.addItem(_NONE_LABEL)
        self.scores_sid_combo.addItems(headers)
        self._populate_target_combo(headers)

        # Set the SID column from the filter
        sid_col = filt.get("student_id_col", "")
        if sid_col:
            idx = self.scores_sid_combo.findText(sid_col)
            if idx >= 0:
                self.scores_sid_combo.setCurrentIndex(idx)

        self.scores_status.setText(f"Filter \"{name}\" applied.")

    def _on_scores_lms_selected(self):
        """Auto-populate SID and target combos when a file is chosen.

        Uses default parsing settings (skip 0, comma delimiter for CSV).
        The teacher can refine by applying a saved filter afterward.
        """
        lms_path = self.scores_lms_file.path()
        if not lms_path or not Path(lms_path).is_file():
            return

        ext = Path(lms_path).suffix.lower()
        try:
            if ext in (".xlsx", ".xls"):
                headers, _ = self._read_excel(lms_path, 0)
            else:
                headers, _ = self._read_text(lms_path, ",", 0)
        except Exception:
            return

        # Populate the SID combo with headers
        self.scores_sid_combo.clear()
        self.scores_sid_combo.addItem(_NONE_LABEL)
        self.scores_sid_combo.addItems(headers)

        # Populate the target column combo
        self._populate_target_combo(headers)

    def _populate_target_combo(self, headers: list[str]):
        """Fill the target-column combo: 'Add new' + existing headers."""
        self.scores_target_combo.blockSignals(True)
        self.scores_target_combo.clear()
        self.scores_target_combo.addItem(_ADD_NEW_COL)
        for h in headers:
            self.scores_target_combo.addItem(h)
        self.scores_target_combo.blockSignals(False)

    def _on_target_col_changed(self, index: int):
        """When 'Add a new column' is selected, prompt for the name."""
        if self.scores_target_combo.currentText() != _ADD_NEW_COL:
            return

        name, ok = QInputDialog.getText(
            self, "New Column",
            "Enter a name for the new score column:",
            QLineEdit.EchoMode.Normal,
            "Exam 1",
        )
        if ok and name.strip():
            # Insert the custom name right after 'Add new' and select it
            self.scores_target_combo.blockSignals(True)
            self.scores_target_combo.insertItem(1, name.strip())
            self.scores_target_combo.setCurrentIndex(1)
            self.scores_target_combo.blockSignals(False)

    @staticmethod
    def _count_orphans_in_report(xlsx_path: str) -> int:
        """Count rows with orphan IDs in an exam report.

        Checks the 'Class Scores' sheet for a FlagDetails column containing
        'ID:orphan'.  Returns 0 if no such column exists.
        """
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            sheet_name = None
            for name in wb.sheetnames:
                if name.lower().replace(" ", "") == "classscores":
                    sheet_name = name
                    break
            if sheet_name is None:
                wb.close()
                return 0

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(rows) < 2:
                return 0

            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            fd_idx = -1
            for i, h in enumerate(headers):
                if h in ("flagdetails", "flag_details", "flag details"):
                    fd_idx = i
                    break
            if fd_idx < 0:
                return 0

            count = 0
            for row in rows[1:]:
                if fd_idx < len(row) and row[fd_idx]:
                    if "ID:orphan" in str(row[fd_idx]):
                        count += 1
            return count
        except Exception:
            return 0

    def _run_write_scores(self):
        """Write MarkShark scores into the LMS gradebook."""
        lms_path = self.scores_lms_file.path()
        results_path = self.scores_results.path()
        output_path = self.scores_output.path()

        if not lms_path or not Path(lms_path).is_file():
            QMessageBox.warning(self, "Missing Input", "Please select the LMS gradebook file.")
            return
        if not results_path or not Path(results_path).is_file():
            QMessageBox.warning(self, "Missing Input", "Please select the MarkShark exam report (exam_report.xlsx).")
            return
        if not output_path:
            QMessageBox.warning(self, "Missing Output", "Please specify an output file path.")
            return

        sid_col = self.scores_sid_combo.currentText()
        if not sid_col or sid_col == _NONE_LABEL:
            QMessageBox.warning(
                self, "No Student ID Column",
                "Please set or load a filter for the Student ID column in the LMS file.",
            )
            return

        target_col = self.scores_target_combo.currentText().strip()
        if not target_col or target_col == _ADD_NEW_COL:
            QMessageBox.warning(
                self, "No Target Column",
                "Please choose an existing column or add a new one for the scores.",
            )
            return

        value_type = self.scores_value_combo.currentText()

        # Warn about unresolved orphan IDs
        orphan_count = self._count_orphans_in_report(results_path)
        if orphan_count > 0:
            reply = QMessageBox.warning(
                self,
                "Orphan Students Detected",
                f"{orphan_count} orphan student(s) found in the exam report.\n\n"
                "Their scores won't match any LMS gradebook entry and will be "
                "skipped during write-back.\n\n"
                "Consider correcting orphan IDs in Review & Correct first.\n\n"
                "Continue anyway?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        # Parse MarkShark exam report
        try:
            score_map = self._parse_markshark_report(results_path, value_type)
        except Exception as e:
            QMessageBox.critical(self, "Report Error", f"Failed to parse exam report:\n{e}")
            return

        if not score_map:
            QMessageBox.warning(self, "No Scores", "No student scores found in the exam report.")
            return

        # Determine output format from extension
        out_ext = Path(output_path).suffix.lower()

        self.scores_status.setText("Writing scores...")
        self.scores_run_btn.setEnabled(False)
        try:
            # Get filter settings for reading the LMS file
            filter_name = self.scores_filter_combo.currentText()
            skip = 0
            delim = ","
            if filter_name and filter_name != "(none — set manually)":
                filt = self._registry.get_filter(filter_name)
                if filt:
                    skip = filt.get("skip_rows", 0)
                    delim = filt.get("delimiter", ",")

            lms_ext = Path(lms_path).suffix.lower()

            if lms_ext in (".xlsx", ".xls"):
                headers, rows = self._read_full_excel(lms_path, skip)
            else:
                headers, rows = self._read_full_text(lms_path, delim, skip)

            if not headers:
                QMessageBox.warning(self, "Empty File", "The LMS file appears to be empty.")
                return

            # Find SID column index
            sid_idx = headers.index(sid_col) if sid_col in headers else -1
            if sid_idx < 0:
                QMessageBox.warning(self, "Column Not Found", f"Column \"{sid_col}\" not found in the LMS file.")
                return

            # Find or create target column
            if target_col in headers:
                target_idx = headers.index(target_col)
            else:
                headers.append(target_col)
                target_idx = len(headers) - 1
                for row in rows:
                    row.append("")

            # Determine absent-student fill value
            fill_absent_zero = self.absent_combo.currentIndex() == 1

            # Write scores — only touch cells where we have a match.
            # Pre-existing data in non-matching rows is preserved so
            # teachers can safely write into a column that already has
            # other scores or notes.
            matched = 0
            for row in rows:
                sid = row[sid_idx].strip() if sid_idx < len(row) else ""
                while len(row) <= target_idx:
                    row.append("")
                if sid in score_map:
                    row[target_idx] = score_map[sid]
                    matched += 1
                elif fill_absent_zero and not row[target_idx].strip():
                    # Only fill blank cells with "0" for absent students;
                    # never overwrite existing content.
                    row[target_idx] = "0"

            # Save output
            if out_ext == ".xlsx":
                self._write_xlsx(output_path, headers, rows)
            else:
                self._write_csv(output_path, headers, rows, delim)

            self.scores_status.setText(
                f"Done — {matched}/{len(score_map)} students matched."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to write scores:\n{e}")
            self.scores_status.setText("Error.")
        finally:
            self.scores_run_btn.setEnabled(True)

    # ------------------------------------------------------------------
    # Score helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_markshark_report(xlsx_path: str, value_type: str) -> dict:
        """Parse exam_report.xlsx 'Class Scores' sheet.

        Returns {student_id: score_value} dict.  The report already has
        corrections applied, so these are final grades.
        """
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # Find the Class Scores sheet (case-insensitive)
        sheet_name = None
        for name in wb.sheetnames:
            if name.lower().replace(" ", "") == "classscores":
                sheet_name = name
                break
        if sheet_name is None:
            wb.close()
            raise ValueError(
                "Could not find a 'Class Scores' sheet in the report. "
                "Make sure you selected an exam_report.xlsx generated by MarkShark."
            )

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return {}

        # First row is headers
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]

        # Find column indices
        sid_idx = -1
        score_idx = -1
        pct_idx = -1
        for i, h in enumerate(headers):
            if h in ("student id", "studentid", "student_id", "sid", "id"):
                sid_idx = i
            elif h in ("score", "correct", "raw score", "points"):
                score_idx = i
            elif h in ("percent", "percentage", "pct", "%"):
                pct_idx = i

        if sid_idx < 0:
            raise ValueError(
                "Could not find a Student ID column in the Class Scores sheet."
            )

        score_map = {}
        for row in rows[1:]:
            sid = str(row[sid_idx]).strip() if sid_idx < len(row) and row[sid_idx] else ""
            if not sid or sid.upper() == "KEY":
                continue
            score = str(row[score_idx]).strip() if score_idx >= 0 and score_idx < len(row) and row[score_idx] is not None else ""
            pct = str(row[pct_idx]).strip() if pct_idx >= 0 and pct_idx < len(row) and row[pct_idx] is not None else ""

            if "Percent" in value_type:
                score_map[sid] = pct if pct else score
            else:
                score_map[sid] = score if score else pct

        return score_map

    def _read_full_text(self, path: str, delimiter: str, skip: int) -> tuple:
        """Read full CSV/TSV (all rows, not preview-limited)."""
        with open(path, newline="", encoding="utf-8-sig") as f:
            for _ in range(skip):
                next(f, None)
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, [])
            headers = [h.strip() for h in headers]
            rows = [[c.strip() for c in row] for row in reader]
        return headers, rows

    def _read_full_excel(self, path: str, skip: int) -> tuple:
        """Read full XLSX (all rows)."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(c) if c is not None else "" for c in row])
        wb.close()
        if skip >= len(all_rows):
            return [], []
        headers = all_rows[skip]
        data = all_rows[skip + 1:]
        return headers, data

    @staticmethod
    def _write_csv(path: str, headers: list, rows: list, delimiter: str = ","):
        """Write a CSV file."""
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerow(headers)
            writer.writerows(rows)

    @staticmethod
    def _write_xlsx(path: str, headers: list, rows: list):
        """Write an XLSX file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(path)
